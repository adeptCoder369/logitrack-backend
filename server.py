from fastapi import FastAPI, APIRouter, HTTPException, UploadFile, File, Query, Depends
from fastapi.responses import FileResponse, StreamingResponse
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
from pathlib import Path
from pydantic import BaseModel, Field, validator
from typing import List, Optional
import uuid
from datetime import datetime, timezone, timedelta
import shutil
import jwt
import bcrypt
import httpx
import secrets
import random
import io

from auth_utils import get_user_product_ids, get_user_depot_ids, build_product_filter

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# MongoDB connection
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

# Create uploads directory
UPLOAD_DIR = ROOT_DIR / "uploads"
UPLOAD_DIR.mkdir(exist_ok=True)

# JWT Configuration
JWT_SECRET = os.environ.get('JWT_SECRET')
if not JWT_SECRET:
    raise ValueError("JWT_SECRET environment variable is required")
JWT_ALGORITHM = "HS256"

# MSG91 Configuration
MSG91_AUTHKEY = os.environ.get('MSG91_AUTHKEY', '')
MSG91_TEMPLATE_ID = os.environ.get('MSG91_TEMPLATE_ID', '')
MSG91_DLT_TE_ID = os.environ.get('MSG91_DLT_TE_ID', '')
MSG91_SENDER_ID = "INFOET"
OTP_EXPIRY_SECONDS = 120  # 2 minutes
MAX_OTP_ATTEMPTS = 5

# Country codes
COUNTRY_CODES = {
    "IN": {"code": "91", "name": "India", "flag": "🇮🇳"},
    "NP": {"code": "977", "name": "Nepal", "flag": "🇳🇵"},
    "BD": {"code": "880", "name": "Bangladesh", "flag": "🇧🇩"},
    "VN": {"code": "84", "name": "Vietnam", "flag": "🇻🇳"},
    "BT": {"code": "975", "name": "Bhutan", "flag": "🇧🇹"},
    "AE": {"code": "971", "name": "UAE", "flag": "🇦🇪"},
}

app = FastAPI()
api_router = APIRouter(prefix="/api")
security = HTTPBearer()

# ============ OTP MODELS ============

class SendOTPRequest(BaseModel):
    mobile: str
    country_code: str = "91"
    purpose: str = "registration"  # registration, login, reset_password
    
    @validator('mobile')
    def validate_mobile(cls, v):
        # Remove any spaces or special characters
        v = ''.join(filter(str.isdigit, v))
        if len(v) != 10:
            raise ValueError("Mobile number must be exactly 10 digits")
        return v

class VerifyOTPRequest(BaseModel):
    mobile: str
    country_code: str = "91"
    otp_code: str
    purpose: str = "registration"

class RegisterRequest(BaseModel):
    name: str
    mobile: str
    country_code: str = "91"
    password: str
    role: str = "Loader"
    email: Optional[str] = None
    depot_id: Optional[str] = None
    otp_verified: bool = False

class LoginWithPasswordRequest(BaseModel):
    mobile: str
    country_code: str = "91"
    password: str

class LoginWithOTPRequest(BaseModel):
    mobile: str
    country_code: str = "91"

class FirstTimeSetupRequest(BaseModel):
    mobile: str
    country_code: str = "91"
    otp_code: str
    new_password: str

class AdminCreateUserRequest(BaseModel):
    name: str
    mobile: str
    country_code: str = "91"
    role: str
    email: Optional[str] = None
    depot_id: Optional[str] = None
    assigned_products: List[str] = []  # List of product IDs user can access
    assigned_depots: List[str] = []  # List of depot IDs user can access

class ResetPasswordRequest(BaseModel):
    mobile: str
    country_code: str = "91"
    otp_code: str
    new_password: str

class TokenResponse(BaseModel):
    token: str
    user: dict

# ============ AUTH MODELS ============

class UserRegister(BaseModel):
    name: str
    mobile: str
    country_code: str = "91"
    password: str
    role: str
    email: Optional[str] = None
    depot_id: Optional[str] = None

class UserLogin(BaseModel):
    mobile: str
    country_code: str = "91"
    password: str

# ============ ENTITY MODELS ============

# Company Model
class CompanyBase(BaseModel):
    name: str
    trade_name: Optional[str] = None
    location: Optional[str] = None
    city: Optional[str] = None
    district: Optional[str] = None
    state: Optional[str] = None
    country: Optional[str] = "India"
    pin_code: Optional[str] = None
    hsn_code: Optional[str] = None
    product_description: Optional[str] = None
    gst_applicability: Optional[str] = None
    gst_number: Optional[str] = None
    website: Optional[str] = None
    primary_email: Optional[str] = None
    secondary_email: Optional[str] = None
    address: Optional[str] = None
    landmark: Optional[str] = None
    emergency_contact: Optional[str] = None
    whatsapp_number: Optional[str] = None
    telephone: Optional[str] = None
    pan_number: Optional[str] = None
    bank_name: Optional[str] = None
    bank_account_number: Optional[str] = None
    ifsc_code: Optional[str] = None
    contact_person_name: Optional[str] = None
    contact_person_mobile: Optional[str] = None

class CompanyCreate(CompanyBase):
    pass

class Company(CompanyBase):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    added_on: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())
    added_by: Optional[str] = None
    users: List[dict] = []  # List of company users

# Company User Model
class CompanyUserBase(BaseModel):
    name: str
    title: Optional[str] = None
    date_of_birth: Optional[str] = None
    marital_status: Optional[str] = None  # Single, Married
    date_of_anniversary: Optional[str] = None
    mobile_number: Optional[str] = None
    email: Optional[str] = None
    whatsapp_number: Optional[str] = None
    emergency_contact: Optional[str] = None
    address: Optional[str] = None
    city: Optional[str] = None
    district: Optional[str] = None
    state: Optional[str] = None
    pin_code: Optional[str] = None
    country: Optional[str] = "India"  # India, Nepal, Bangladesh, Bhutan, Any Other
    pan_number: Optional[str] = None
    aadhaar_number: Optional[str] = None
    photo_url: Optional[str] = None
    remarks: Optional[str] = None

class CompanyUserCreate(CompanyUserBase):
    pass

class CompanyUser(CompanyUserBase):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    company_id: str
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())

# Transporter Model
class TransporterBase(BaseModel):
    name: str
    trade_name: Optional[str] = None
    contact_person_name: Optional[str] = None
    mobile_number: Optional[str] = None
    email: Optional[str] = None
    address: Optional[str] = None
    gst_number: Optional[str] = None
    industry_type: Optional[str] = None

class TransporterCreate(TransporterBase):
    pass

class Transporter(TransporterBase):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())

# Truck Model
class DriverInfo(BaseModel):
    name: str
    mobile: Optional[str] = None
    is_primary: bool = False

class TruckBase(BaseModel):
    vehicle_number: str
    transporter_id: Optional[str] = None
    transporter_name: Optional[str] = None
    capacity_mt: Optional[float] = None
    tare_weight_mt: Optional[float] = None
    make_model: Optional[str] = None
    driver_name: Optional[str] = None  # Primary/default driver
    driver_mobile: Optional[str] = None
    helper_name: Optional[str] = None
    helper_mobile: Optional[str] = None
    drivers: Optional[List[dict]] = []  # List of all drivers {name, mobile, is_primary}
    current_status: Optional[str] = "Idle"
    front_photo: Optional[str] = None
    back_photo: Optional[str] = None

class TruckCreate(TruckBase):
    pass

class Truck(TruckBase):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())

# Product Model
class ProductBase(BaseModel):
    product_name: str
    product_code: Optional[str] = None
    product_description: Optional[str] = None
    unit_of_measurement: Optional[str] = "MT"
    category: Optional[str] = None
    hsn_code: Optional[str] = None

class ProductCreate(ProductBase):
    pass

class Product(ProductBase):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())

# Depot Model with Inventory Wallet
class DepotBase(BaseModel):
    name: str
    location: Optional[str] = None
    city: Optional[str] = None
    state: Optional[str] = None
    address: Optional[str] = None
    contact_person_name: Optional[str] = None
    contact_mobile: Optional[str] = None
    storage_capacity: Optional[float] = None
    warehouse_type: Optional[str] = None

class DepotCreate(DepotBase):
    pass

class Depot(DepotBase):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())

# Depot Inventory (Wallet) Model
class DepotInventory(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    depot_id: str
    company_id: str
    depot_name: str
    product_id: str
    product_name: str
    product_code: Optional[str] = None
    total_received: float = 0
    total_dispatched: float = 0
    available_quantity: float = 0
    last_updated: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())

# Railway Siding Model
class RailwaySidingBase(BaseModel):
    siding_name: str
    siding_code: str
    location: Optional[str] = None
    station_name: Optional[str] = None
    state: Optional[str] = None
    contact_person_name: Optional[str] = None
    contact_mobile: Optional[str] = None
    remarks: Optional[str] = None

class RailwaySidingCreate(RailwaySidingBase):
    pass

class RailwaySiding(RailwaySidingBase):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())

# Delivery Order (Master Record)
class DeliveryOrderBase(BaseModel):
    transport_mode: str = "Road"  # "Road" or "Railway"
    from_company_id: Optional[str] = None
    from_company_name: Optional[str] = None
    product_id: Optional[str] = None
    product_name: Optional[str] = None
    product_code: Optional[str] = None
    total_quantity_mt: float
    to_depot_id: Optional[str] = None
    to_depot_name: Optional[str] = None
    # Railway-specific fields
    loading_siding_id: Optional[str] = None
    loading_siding_name: Optional[str] = None
    loading_siding_code: Optional[str] = None
    destination_siding_id: Optional[str] = None
    destination_siding_name: Optional[str] = None
    destination_siding_code: Optional[str] = None
    remarks: Optional[str] = None

class DeliveryOrderCreate(DeliveryOrderBase):
    pass

class DeliveryOrder(DeliveryOrderBase):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    do_order_no: str = ""
    do_date: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())
    lifted_quantity_mt: float = 0
    remaining_quantity_mt: float = 0
    status: str = "Open"
    added_by: Optional[str] = None
    added_by_name: Optional[str] = None
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())

# Lifting Model
class LiftingBase(BaseModel):
    lifting_type: str = "Primary"
    transport_mode: str = "Road"  # "Road" or "Railway"
    delivery_order_id: Optional[str] = None
    delivery_order_no: Optional[str] = None
    product_id: Optional[str] = None
    product_name: Optional[str] = None
    product_code: Optional[str] = None
    quantity_mt: float
    loading_point_type: str = "Company"
    loading_point_id: Optional[str] = None
    loading_point_name: Optional[str] = None
    date_of_loading: Optional[str] = None
    time_of_loading: Optional[str] = None
    # Road (Truck) specific fields
    vehicle_id: Optional[str] = None
    vehicle_number: Optional[str] = None
    transporter_name: Optional[str] = None
    driver_name: Optional[str] = None
    driver_mobile: Optional[str] = None
    helper_name: Optional[str] = None
    helper_mobile: Optional[str] = None
    # Railway specific fields
    loading_siding_id: Optional[str] = None
    loading_siding_name: Optional[str] = None
    loading_siding_code: Optional[str] = None
    destination_siding_id: Optional[str] = None
    destination_siding_name: Optional[str] = None
    destination_siding_code: Optional[str] = None
    # Common fields
    tare_weight_mt: Optional[float] = None
    gross_weight_mt: Optional[float] = None
    net_weight_mt: Optional[float] = None
    weight_slip: Optional[str] = None
    unloading_point_type: str = "Depot"
    unloading_point_id: Optional[str] = None
    unloading_point_name: Optional[str] = None

class LiftingCreate(LiftingBase):
    pass

class Lifting(LiftingBase):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    lifting_no: str = ""
    loading_status: str = "Loaded"
    loaded_by: Optional[str] = None
    loaded_by_name: Optional[str] = None
    unloading_status: str = "Pending"
    date_of_unloading: Optional[str] = None
    time_of_unloading: Optional[str] = None
    verified_by: Optional[str] = None
    verified_by_name: Optional[str] = None
    verified_at: Optional[str] = None
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())

class VerifyUnloadRequest(BaseModel):
    date_of_unloading: str
    time_of_unloading: str

# ============ AUTH HELPERS ============

def hash_password(password: str) -> str:
    return bcrypt.hashpw(password.encode(), bcrypt.gensalt()).decode()

def verify_password(password: str, hashed: str) -> bool:
    return bcrypt.checkpw(password.encode(), hashed.encode())

def create_token(user_data: dict) -> str:
    payload = {
        "user_id": user_data["id"],
        "company_id": user_data.get("company_id"),
        "mobile": user_data["mobile"],
        "role": user_data["role"],
        "name": user_data["name"],
        "exp": datetime.now(timezone.utc).timestamp() + 86400 * 7
    }
    return jwt.encode(payload, JWT_SECRET, algorithm=JWT_ALGORITHM)

async def get_current_user(credentials: HTTPAuthorizationCredentials = Depends(security)):
    try:
        token = credentials.credentials
        payload = jwt.decode(token, JWT_SECRET, algorithms=[JWT_ALGORITHM])
        user = await db.users.find_one({"id": payload["user_id"]}, {"_id": 0, "password": 0})
        if not user:
            raise HTTPException(status_code=401, detail="User not found")
        return user
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Token expired")
    except jwt.InvalidTokenError:
        raise HTTPException(status_code=401, detail="Invalid token")

# ============ PERMISSION HELPERS ============

# Define DEFAULT_PERMISSIONS here so it can be used by helper functions
PERMISSION_DEFAULTS = {
    "Dashboard": {"Management": True, "Admin": True, "Loader": True, "Depot Manager": True, "Depot Staff": True},
    "Delivery Orders (View)": {"Management": True, "Admin": True, "Loader": False, "Depot Manager": False, "Depot Staff": False},
    "Delivery Orders (Create)": {"Management": True, "Admin": True, "Loader": False, "Depot Manager": False, "Depot Staff": False},
    "Primary Liftings (Create)": {"Management": True, "Admin": True, "Loader": True, "Depot Manager": False, "Depot Staff": False},
    "Primary Liftings (View)": {
        "Management": True, 
        "Admin": True,
        "Loader": True,
        "Depot Manager": False,
        "Depot Staff": False
    },
    "Primary Liftings (Update)": {
        "Management": True, 
        "Admin": True,
        "Loader": False,
        "Depot Manager": False,
        "Depot Staff": False
    },
    "Primary Liftings (Delete)": {
        "Management": True, 
        "Admin": True,
        "Loader": False,
        "Depot Manager": False,
        "Depot Staff": False
    },
    "Secondary Liftings (Create)": {"Management": True, "Admin": True, "Loader": False, "Depot Manager": True, "Depot Staff": True},
    "Secondary Liftings (View)": {
        "Management": True, 
        "Admin": True,
        "Loader": False,
        "Depot Manager": True,
        "Depot Staff": True
    },
    "Secondary Liftings (Update)": {
        "Management": True, 
        "Admin": True,
        "Loader": False,
        "Depot Manager": True,
        "Depot Staff": False
    },
    "Secondary Liftings (Delete)": {
        "Management": True, 
        "Admin": True,
        "Loader": False,
        "Depot Manager": False,
        "Depot Staff": False
    },
    "Verification (Unloading)": {"Management": True, "Admin": True, "Loader": False, "Depot Manager": True, "Depot Staff": False},
    "Inventory Wallet": {"Management": True, "Admin": True, "Loader": False, "Depot Manager": True, "Depot Staff": True},
    "Inventory Wallet (View)": {
        "Management": True, 
        "Admin": True,
        "Loader": False,
        "Depot Manager": True,
        "Depot Staff": True
    },
    "Inventory Wallet (Update)": {
        "Management": True, 
        "Admin": True,
        "Loader": False,
        "Depot Manager": True,
        "Depot Staff": False
    },
    "Inventory Wallet (Delete)": {
        "Management": True, 
        "Admin": True,
        "Loader": False,
        "Depot Manager": False,
        "Depot Staff": False
    },
    "DO Wallet": {"Management": True, "Admin": True, "Loader": True, "Depot Manager": False, "Depot Staff": False},
    "DO Wallet (View)": {
        "Management": True, 
        "Admin": True,
        "Loader": True,
        "Depot Manager": False,
        "Depot Staff": False
    },
    "DO Wallet (Update)": {
        "Management": True, 
        "Admin": True,
        "Loader": False,
        "Depot Manager": False,
        "Depot Staff": False
    },
    "Company Reports": {"Management": True, "Admin": True, "Loader": False, "Depot Manager": True, "Depot Staff": True},
    "Companies": {"Management": True, "Admin": True, "Loader": False, "Depot Manager": False, "Depot Staff": False},
    "Companies (View)": {
        "Management": True, 
        "Admin": True,
        "Loader": False,
        "Depot Manager": False,
        "Depot Staff": False
    },
    "Companies (Create)": {
        "Management": True, 
        "Admin": True,
        "Loader": False,
        "Depot Manager": False,
        "Depot Staff": False
    },
    "Companies (Update)": {
        "Management": True, 
        "Admin": True,
        "Loader": False,
        "Depot Manager": False,
        "Depot Staff": False
    },
    "Companies (Delete)": {
        "Management": True, 
        "Admin": True,
        "Loader": False,
        "Depot Manager": False,
        "Depot Staff": False
    },
    "Transporters": {"Management": True, "Admin": True, "Loader": False, "Depot Manager": False, "Depot Staff": False},
    "Trucks": {"Management": True, "Admin": True, "Loader": True, "Depot Manager": False, "Depot Staff": False},
    "Trucks (View)": {
        "Management": True, 
        "Admin": True,
        "Loader": True,
        "Depot Manager": False,
        "Depot Staff": False
    },
    "Trucks (Create)": {
        "Management": True, 
        "Admin": True,
        "Loader": False,
        "Depot Manager": False,
        "Depot Staff": False
    },
    "Trucks (Update)": {
        "Management": True, 
        "Admin": True,
        "Loader": False,
        "Depot Manager": False,
        "Depot Staff": False
    },
    "Trucks (Delete)": {
        "Management": True, 
        "Admin": True,
        "Loader": False,
        "Depot Manager": False,
        "Depot Staff": False
    },
    "Railway Sidings (View)": {
        "Management": True, 
        "Admin": True,
        "Loader": True,
        "Depot Manager": False,
        "Depot Staff": False
    },
    "Railway Sidings (Create)": {
        "Management": True, 
        "Admin": True,
        "Loader": False,
        "Depot Manager": False,
        "Depot Staff": False
    },
    "Railway Sidings (Update)": {
        "Management": True, 
        "Admin": True,
        "Loader": False,
        "Depot Manager": False,
        "Depot Staff": False
    },
    "Railway Sidings (Delete)": {
        "Management": True, 
        "Admin": True,
        "Loader": False,
        "Depot Manager": False,
        "Depot Staff": False
    },
    "Products": {"Management": True, "Admin": True, "Loader": False, "Depot Manager": False, "Depot Staff": False},
    "Depots": {"Management": True, "Admin": True, "Loader": False, "Depot Manager": False, "Depot Staff": False},
    "User Management": {"Management": True, "Admin": True, "Loader": False, "Depot Manager": False, "Depot Staff": False},
    "Role Permissions": {"Management": True, "Admin": True, "Loader": False, "Depot Manager": False, "Depot Staff": False},
    "Analytics": {"Management": True, "Admin": True, "Loader": False, "Depot Manager": False, "Depot Staff": False},
    "Liftings": {"Management": True, "Admin": True, "Loader": True, "Depot Manager": True, "Depot Staff": True},
}
PERMISSION_DEFAULTS.update({
    "Delivery Orders (Update)": {
        "Management": True, 
        "Admin": True,
        "Loader": False,
        "Depot Manager": False,
        "Depot Staff": False
    },
    "Delivery Orders (Delete)": {
        "Management": True, 
        "Admin": True,
        "Loader": False,
        "Depot Manager": False,
        "Depot Staff": False
    }
})
PERMISSION_DEFAULTS.update({
    "Liftings (View)": {
        "Management": True, 
        "Admin": True,
        "Loader": True,
        "Depot Manager": True,
        "Depot Staff": True
    },
    "Liftings (Update)": {
        "Management": True, 
        "Admin": True,
        "Loader": False,
        "Depot Manager": True,
        "Depot Staff": False
    },
    "Liftings (Delete)": {
        "Management": True, 
        "Admin": True,
        "Loader": False,
        "Depot Manager": False,
        "Depot Staff": False
    }
})
PERMISSION_DEFAULTS.update({
    "Inventory Wallet (View)": {
        "Management": True, 
        "Admin": True,
        "Loader": False,
        "Depot Manager": True,
        "Depot Staff": True
    },
    "Inventory Wallet (Update)": {
        "Management": True, 
        "Admin": True,
        "Loader": False,
        "Depot Manager": True,
        "Depot Staff": False
    },
    "Inventory Wallet (Delete)": {
        "Management": True, 
        "Admin": True,
        "Loader": False,
        "Depot Manager": False,
        "Depot Staff": False
    }
})
PERMISSION_DEFAULTS.update({
    "Companies (View)": {
        "Management": True, 
        "Admin": True,
        "Loader": False,
        "Depot Manager": False,
        "Depot Staff": False
    },
    "Companies (Update)": {
        "Management": True, 
        "Admin": True,
        "Loader": False,
        "Depot Manager": False,
        "Depot Staff": False
    },
    "Companies (Delete)": {
        "Management": True, 
        "Admin": True,
        "Loader": False,
        "Depot Manager": False,
        "Depot Staff": False
    }
})
PERMISSION_DEFAULTS.update({
    "Transporters (View)": {
        "Management": True, 
        "Admin": True,
        "Loader": False,
        "Depot Manager": False,
        "Depot Staff": False
    },
    "Transporters (Create)": {
        "Management": True, 
        "Admin": True,
        "Loader": False,
        "Depot Manager": False,
        "Depot Staff": False
    },
    "Transporters (Update)": {
        "Management": True, 
        "Admin": True,
        "Loader": False,
        "Depot Manager": False,
        "Depot Staff": False
    },
    "Transporters (Delete)": {
        "Management": True, 
        "Admin": True,
        "Loader": False,
        "Depot Manager": False,
        "Depot Staff": False
    }
})
PERMISSION_DEFAULTS.update({
    "Trucks (View)": {
        "Management": True, 
        "Admin": True,
        "Loader": True,
        "Depot Manager": False,
        "Depot Staff": False
    },
    "Trucks (Create)": {
        "Management": True, 
        "Admin": True,
        "Loader": False,
        "Depot Manager": False,
        "Depot Staff": False
    },
    "Trucks (Update)": {
        "Management": True, 
        "Admin": True,
        "Loader": False,
        "Depot Manager": False,
        "Depot Staff": False
    },
    "Trucks (Delete)": {
        "Management": True, 
        "Admin": True,
        "Loader": False,
        "Depot Manager": False,
        "Depot Staff": False
    }
})
PERMISSION_DEFAULTS.update({
    "Products (View)": {
        "Management": True, 
        "Admin": True,
        "Loader": False,
        "Depot Manager": False,
        "Depot Staff": False
    },
    "Products (Create)": {
        "Management": True, 
        "Admin": True,
        "Loader": False,
        "Depot Manager": False,
        "Depot Staff": False
    },
    "Products (Update)": {
        "Management": True, 
        "Admin": True,
        "Loader": False,
        "Depot Manager": False,
        "Depot Staff": False
    },
    "Products (Delete)": {
        "Management": True, 
        "Admin": True,
        "Loader": False,
        "Depot Manager": False,
        "Depot Staff": False
    }
})
PERMISSION_DEFAULTS.update({
    "Depots (View)": {
        "Management": True, 
        "Admin": True,
        "Loader": False,
        "Depot Manager": False,
        "Depot Staff": False
    },
    "Depots (Create)": {
        "Management": True, 
        "Admin": True,
        "Loader": False,
        "Depot Manager": False,
        "Depot Staff": False
    },
    "Depots (Update)": {
        "Management": True, 
        "Admin": True,
        "Loader": False,
        "Depot Manager": False,
        "Depot Staff": False
    },
    "Depots (Delete)": {
        "Management": True, 
        "Admin": True,
        "Loader": False,
        "Depot Manager": False,
        "Depot Staff": False
    }
})
PERMISSION_DEFAULTS.update({
    "User Management (View)": {
        "Management": True, 
        "Admin": True,
        "Loader": False,
        "Depot Manager": False,
        "Depot Staff": False
    },
    "User Management (Create)": {
        "Management": True, 
        "Admin": True,
        "Loader": False,
        "Depot Manager": False,
        "Depot Staff": False
    },
    "User Management (Update)": {
        "Management": True, 
        "Admin": True,
        "Loader": False,
        "Depot Manager": False,
        "Depot Staff": False
    },
    "User Management (Delete)": {
        "Management": True, 
        "Admin": True,
        "Loader": False,
        "Depot Manager": False,
        "Depot Staff": False
    }
})
PERMISSION_DEFAULTS.update({
    "Purchase Orders (View)": {
        "Management": True, 
        "Admin": True,
        "Loader": False,
        "Depot Manager": False,
        "Depot Staff": False
    },
    "Purchase Orders (Create)": {
        "Management": True, 
        "Admin": True,
        "Loader": False,
        "Depot Manager": False,
        "Depot Staff": False
    },
    "Purchase Orders (Update)": {
        "Management": True, 
        "Admin": True,
        "Loader": False,
        "Depot Manager": False,
        "Depot Staff": False
    },
    "Purchase Orders (Delete)": {
        "Management": True, 
        "Admin": True,
        "Loader": False,
        "Depot Manager": False,
        "Depot Staff": False
    }
})
PERMISSION_DEFAULTS.update({
    "Schedule Pickup": {
        "Management": True, 
        "Admin": True,
        "Loader": False,
        "Depot Manager": True,
        "Depot Supervisor": True,
        "Depot Staff": False
    },
    "Pickup (Execution)": {
        "Management": True, 
        "Admin": True,
        "Loader": True,
        "Depot Manager": True,
        "Depot Supervisor": True,
        "Depot Staff": False
    },
    "Verify Pickup": {
        "Management": True, 
        "Admin": True,
        "Loader": False,
        "Depot Manager": True,
        "Depot Supervisor": True,
        "Depot Staff": False
    }
})



async def fetch_permissions():
    """Get current permissions from database"""
    perm_doc = await db.permissions.find_one({"id": "role_permissions"}, {"_id": 0})
    if not perm_doc:
        return PERMISSION_DEFAULTS
    return perm_doc.get("permissions", PERMISSION_DEFAULTS)

async def check_permission(user: dict, permission_key: str):
    """Check if user has a specific permission"""
    if not user:
        raise HTTPException(status_code=401, detail="Not authenticated")
    
    # Management always has all permissions
    if user.get("role") == "Management":
        return True
    
    permissions = await fetch_permissions()
    has_permission = permissions.get(permission_key, {}).get(user.get("role"), False)
    
    if not has_permission:
        raise HTTPException(
            status_code=403, 
            detail=f"You don't have permission for: {permission_key}"
        )
    return True

def require_permission(permission_key: str):
    """Dependency factory for permission checking"""
    async def permission_checker(current_user: dict = Depends(get_current_user)):
        await check_permission(current_user, permission_key)
        return current_user
    return permission_checker

# ============ OTP HELPERS ============

def generate_otp(length: int = 6) -> str:
    return ''.join([str(random.randint(0, 9)) for _ in range(length)])

async def send_otp_via_msg91(mobile: str, country_code: str, otp_code: str) -> dict:
    """Send OTP via MSG91 SMS API"""
    full_mobile = f"{country_code}{mobile}"
    expiry_mins = OTP_EXPIRY_SECONDS // 60 or 1
    
    # Use MSG91 SMS API (not OTP API) with DLT template
    url = "https://api.msg91.com/api/v2/sendsms"
    
    headers = {
        "authkey": MSG91_AUTHKEY,
        "Content-Type": "application/json"
    }
    
    # Construct message with OTP and expiry
    message = f"{otp_code} is the OTP for accessing your infoEIGHT account. Please do not share it with anyone. OTP will be valid for {expiry_mins} mins."
    
    payload = {
        "sender": MSG91_SENDER_ID,
        "route": "4",  # Transactional route
        "country": country_code,
        "DLT_TE_ID": MSG91_DLT_TE_ID,
        "sms": [
            {
                "message": message,
                "to": [mobile]
            }
        ]
    }
    
    logging.info(f"Sending OTP to {full_mobile} via MSG91 SMS API")
    
    async with httpx.AsyncClient() as client:
        try:
            response = await client.post(url, json=payload, headers=headers, timeout=10)
            result = response.json()
            logging.info(f"MSG91 Response: {result}")
            return result
        except Exception as e:
            logging.error(f"MSG91 API Error: {str(e)}")
            # Return success for demo even if MSG91 fails
            return {"type": "success", "message": "OTP sent (demo mode)"}

# ============ OTP ROUTES ============

@api_router.get("/country-codes")
async def get_country_codes():
    """Get list of supported country codes"""
    return list(COUNTRY_CODES.values())

@api_router.post("/otp/send")
async def send_otp(request: SendOTPRequest):
    """Send OTP to mobile number"""
    full_mobile = f"{request.country_code}{request.mobile}"
    
    # Check rate limiting - max 3 OTPs per 5 minutes
    five_minutes_ago = (datetime.now(timezone.utc) - timedelta(minutes=5)).isoformat()
    recent_otps = await db.otps.count_documents({
        "mobile": full_mobile,
        "created_at": {"$gte": five_minutes_ago}
    })
    
    if recent_otps >= 3:
        raise HTTPException(status_code=429, detail="Too many OTP requests. Please wait 5 minutes.")
    
    # Generate OTP
    otp_code = generate_otp()
    
    # Store OTP in database
    expires_at = datetime.now(timezone.utc) + timedelta(seconds=OTP_EXPIRY_SECONDS)
    otp_doc = {
        "id": str(uuid.uuid4()),
        "mobile": full_mobile,
        "country_code": request.country_code,
        "otp_code": otp_code,
        "purpose": request.purpose,
        "verified": False,
        "attempts": 0,
        "created_at": datetime.now(timezone.utc).isoformat(),
        "expires_at": expires_at.isoformat()
    }
    
    await db.otps.insert_one(otp_doc)
    
    # Send OTP via MSG91
    result = await send_otp_via_msg91(request.mobile, request.country_code, otp_code)
    
    return {
        "success": True,
        "message": f"OTP sent to +{full_mobile}",
        "expires_in_seconds": OTP_EXPIRY_SECONDS,
        # For testing - remove in production
        "demo_otp": otp_code
    }

@api_router.post("/otp/verify")
async def verify_otp(request: VerifyOTPRequest):
    """Verify OTP"""
    full_mobile = f"{request.country_code}{request.mobile}"
    
    # Find latest unverified OTP for this mobile and purpose
    otp_record = await db.otps.find_one(
        {
            "mobile": full_mobile,
            "purpose": request.purpose,
            "verified": False
        },
        sort=[("created_at", -1)]
    )
    
    if not otp_record:
        raise HTTPException(status_code=404, detail="No pending OTP found. Please request a new one.")
    
    # Check expiry - parse ISO string to datetime
    expires_at = datetime.fromisoformat(otp_record["expires_at"].replace('Z', '+00:00'))
    if datetime.now(timezone.utc) > expires_at:
        raise HTTPException(status_code=410, detail="OTP has expired. Please request a new one.")
    
    # Check attempts
    if otp_record["attempts"] >= MAX_OTP_ATTEMPTS:
        raise HTTPException(status_code=429, detail="Too many failed attempts. Please request a new OTP.")
    
    # Verify OTP
    if otp_record["otp_code"] != request.otp_code:
        await db.otps.update_one(
            {"id": otp_record["id"]},
            {"$inc": {"attempts": 1}}
        )
        remaining = MAX_OTP_ATTEMPTS - otp_record["attempts"] - 1
        raise HTTPException(status_code=401, detail=f"Invalid OTP. {remaining} attempts remaining.")
    
    # Mark as verified
    await db.otps.update_one(
        {"id": otp_record["id"]},
        {"$set": {"verified": True}}
    )
    
    # Generate verification token
    verification_token = secrets.token_urlsafe(32)
    
    return {
        "success": True,
        "message": "OTP verified successfully",
        "verification_token": verification_token
    }

@api_router.post("/otp/resend")
async def resend_otp(request: SendOTPRequest):
    """Resend OTP"""
    return await send_otp(request)

# ============ AUTH ROUTES ============

@api_router.post("/auth/register", response_model=TokenResponse)
async def register(data: UserRegister):
    """Register new user with OTP verification"""
    full_mobile = f"{data.country_code}{data.mobile}"
    
    # Check if mobile already exists
    existing = await db.users.find_one({"mobile": full_mobile})
    if existing:
        raise HTTPException(status_code=400, detail="Mobile number already registered")
    
    # Verify OTP was verified for this mobile
    verified_otp = await db.otps.find_one({
        "mobile": full_mobile,
        "purpose": "registration",
        "verified": True
    }, sort=[("created_at", -1)])
    
    if not verified_otp:
        raise HTTPException(status_code=400, detail="Please verify your mobile number first")
    
    user_id = str(uuid.uuid4())
    user_doc = {
        "id": user_id,
        "name": data.name,
        "mobile": full_mobile,
        "country_code": data.country_code,
        "password": hash_password(data.password),
        "role": data.role,
        "email": data.email,
        "depot_id": data.depot_id,
        "otp_verified": True,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.users.insert_one(user_doc)
    
    # Clean up used OTPs
    await db.otps.delete_many({"mobile": full_mobile, "purpose": "registration"})
    
    user_response = {k: v for k, v in user_doc.items() if k != "password" and k != "_id"}
    token = create_token(user_doc)
    return {"token": token, "user": user_response}

@api_router.post("/auth/login")
async def login(data: UserLogin):
    """Login with mobile and password. Detects first-time users."""
    full_mobile = f"{data.country_code}{data.mobile}"
    
    user = await db.users.find_one({"mobile": full_mobile})
    if not user:
        raise HTTPException(status_code=401, detail="Invalid mobile number or password")
    
    # Check if this is a first-time user (password not set)
    if not user.get("password_set", True):
        # First-time login - send OTP for password setup
        otp_code = generate_otp()
        expires_at = datetime.now(timezone.utc) + timedelta(seconds=OTP_EXPIRY_SECONDS)
        otp_doc = {
            "id": str(uuid.uuid4()),
            "mobile": full_mobile,
            "country_code": data.country_code,
            "otp_code": otp_code,
            "purpose": "first_time_setup",
            "verified": False,
            "attempts": 0,
            "created_at": datetime.now(timezone.utc).isoformat(),
            "expires_at": expires_at.isoformat()
        }
        await db.otps.insert_one(otp_doc)
        
        # Send OTP via MSG91
        await send_otp_via_msg91(data.mobile, data.country_code, otp_code)
        
        return {
            "first_time_login": True,
            "message": f"OTP sent to +{full_mobile}. Please set your password.",
            "expires_in_seconds": OTP_EXPIRY_SECONDS,
            "demo_otp": otp_code  # Remove in production
        }
    
    # Regular login with password
    if not verify_password(data.password, user["password"]):
        raise HTTPException(status_code=401, detail="Invalid mobile number or password")
    
    user_response = {k: v for k, v in user.items() if k != "password" and k != "_id"}
    token = create_token(user)
    return {"token": token, "user": user_response}

@api_router.post("/auth/first-time-setup", response_model=TokenResponse)
async def first_time_setup(data: FirstTimeSetupRequest):
    """Complete first-time login by verifying OTP and setting password"""
    full_mobile = f"{data.country_code}{data.mobile}"
    
    # Find user
    user = await db.users.find_one({"mobile": full_mobile})
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    
    if user.get("password_set", True):
        raise HTTPException(status_code=400, detail="Password already set. Please use regular login.")
    
    # Verify OTP
    otp_record = await db.otps.find_one(
        {
            "mobile": full_mobile,
            "purpose": "first_time_setup",
            "verified": False
        },
        sort=[("created_at", -1)]
    )
    
    if not otp_record:
        raise HTTPException(status_code=404, detail="No pending OTP found. Please try logging in again.")
    
    expires_at = datetime.fromisoformat(otp_record["expires_at"].replace('Z', '+00:00'))
    if datetime.now(timezone.utc) > expires_at:
        raise HTTPException(status_code=410, detail="OTP has expired. Please try logging in again.")
    
    if otp_record["otp_code"] != data.otp_code:
        await db.otps.update_one({"id": otp_record["id"]}, {"$inc": {"attempts": 1}})
        raise HTTPException(status_code=401, detail="Invalid OTP")
    
    # Set password and mark as verified
    await db.users.update_one(
        {"mobile": full_mobile},
        {
            "$set": {
                "password": hash_password(data.new_password),
                "password_set": True,
                "otp_verified": True
            }
        }
    )
    
    # Clean up OTPs
    await db.otps.delete_many({"mobile": full_mobile, "purpose": "first_time_setup"})
    
    # Get updated user and create token
    user = await db.users.find_one({"mobile": full_mobile})
    user_response = {k: v for k, v in user.items() if k != "password" and k != "_id"}
    token = create_token(user)
    
    return {"token": token, "user": user_response}

@api_router.post("/auth/login-otp")
async def login_with_otp(data: LoginWithOTPRequest):
    """Request OTP for login"""
    full_mobile = f"{data.country_code}{data.mobile}"
    
    # Check if user exists
    user = await db.users.find_one({"mobile": full_mobile})
    if not user:
        raise HTTPException(status_code=404, detail="No account found with this mobile number")
    
    # Send OTP
    otp_request = SendOTPRequest(
        mobile=data.mobile,
        country_code=data.country_code,
        purpose="login"
    )
    return await send_otp(otp_request)

@api_router.post("/auth/login-otp/verify", response_model=TokenResponse)
async def verify_login_otp(request: VerifyOTPRequest):
    """Verify OTP and login"""
    full_mobile = f"{request.country_code}{request.mobile}"
    
    # Verify OTP
    otp_record = await db.otps.find_one(
        {
            "mobile": full_mobile,
            "purpose": "login",
            "verified": False
        },
        sort=[("created_at", -1)]
    )
    
    if not otp_record:
        raise HTTPException(status_code=404, detail="No pending OTP found")
    
    expires_at = datetime.fromisoformat(otp_record["expires_at"].replace('Z', '+00:00'))
    if datetime.now(timezone.utc) > expires_at:
        raise HTTPException(status_code=410, detail="OTP has expired")
    
    if otp_record["otp_code"] != request.otp_code:
        await db.otps.update_one({"id": otp_record["id"]}, {"$inc": {"attempts": 1}})
        raise HTTPException(status_code=401, detail="Invalid OTP")
    
    # Mark verified
    await db.otps.update_one({"id": otp_record["id"]}, {"$set": {"verified": True}})
    
    # Get user and login
    user = await db.users.find_one({"mobile": full_mobile})
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    
    user_response = {k: v for k, v in user.items() if k != "password" and k != "_id"}
    token = create_token(user)
    
    # Clean up OTPs
    await db.otps.delete_many({"mobile": full_mobile, "purpose": "login"})
    
    return {"token": token, "user": user_response}

@api_router.post("/auth/forgot-password")
async def forgot_password(data: LoginWithOTPRequest):
    """Request OTP for password reset"""
    full_mobile = f"{data.country_code}{data.mobile}"
    
    # Check if user exists
    user = await db.users.find_one({"mobile": full_mobile})
    if not user:
        raise HTTPException(status_code=404, detail="No account found with this mobile number")
    
    # Send OTP
    otp_request = SendOTPRequest(
        mobile=data.mobile,
        country_code=data.country_code,
        purpose="reset_password"
    )
    return await send_otp(otp_request)

@api_router.post("/auth/reset-password")
async def reset_password(data: ResetPasswordRequest):
    """Reset password with OTP verification"""
    full_mobile = f"{data.country_code}{data.mobile}"
    
    # Verify OTP
    otp_record = await db.otps.find_one(
        {
            "mobile": full_mobile,
            "purpose": "reset_password",
            "verified": False
        },
        sort=[("created_at", -1)]
    )
    
    if not otp_record:
        raise HTTPException(status_code=404, detail="No pending OTP found. Request password reset first.")
    
    expires_at = datetime.fromisoformat(otp_record["expires_at"].replace('Z', '+00:00'))
    if datetime.now(timezone.utc) > expires_at:
        raise HTTPException(status_code=410, detail="OTP has expired")
    
    if otp_record["otp_code"] != data.otp_code:
        await db.otps.update_one({"id": otp_record["id"]}, {"$inc": {"attempts": 1}})
        raise HTTPException(status_code=401, detail="Invalid OTP")
    
    # Update password
    await db.users.update_one(
        {"mobile": full_mobile},
        {"$set": {"password": hash_password(data.new_password)}}
    )
    
    # Clean up OTPs
    await db.otps.delete_many({"mobile": full_mobile, "purpose": "reset_password"})
    
    return {"success": True, "message": "Password reset successfully. Please login with your new password."}

@api_router.get("/auth/me")
async def get_me(current_user: dict = Depends(get_current_user)):
    return current_user

# ============ ADMIN USER MANAGEMENT ============

@api_router.post("/admin/users")
async def admin_create_user(data: AdminCreateUserRequest, current_user: dict = Depends(get_current_user)):
    """Management-only endpoint to create new users"""
    # Check if current user is Management
    if current_user.get("role") != "Management":
        raise HTTPException(status_code=403, detail="Only Management can create users")
    
    full_mobile = f"{data.country_code}{data.mobile}"
    
    # Check if mobile already exists
    existing = await db.users.find_one({"mobile": full_mobile})
    if existing:
        raise HTTPException(status_code=400, detail="Mobile number already registered")
    
    user_id = str(uuid.uuid4())
    user_doc = {
        "id": user_id,
        "name": data.name,
        "mobile": full_mobile,
        "country_code": data.country_code,
        "password": "",  # Empty password - user will set on first login
        "password_set": False,  # Flag to indicate first-time user
        "role": data.role,
        "email": data.email,
        "depot_id": data.depot_id,
        "company_id": current_user.get("company_id"),
        "assigned_products": data.assigned_products,
        "assigned_depots": data.assigned_depots,
        "excluded_products": [],
        "excluded_depots": [],
        "otp_verified": False,
        "created_by": current_user["id"],
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    if not current_user.get("company_id") and not current_user.get("is_master_admin"):
        raise HTTPException(400, "User must belong to a company")
    await db.users.insert_one(user_doc)
    
    user_response = {k: v for k, v in user_doc.items() if k != "password" and k != "_id"}
    return {"success": True, "message": f"User created. They can login with mobile +{full_mobile}", "user": user_response}

# ============ FILE UPLOAD ============

@api_router.post("/upload")
async def upload_file(file: UploadFile = File(...)):
    # Preserve original filename with a unique prefix to avoid conflicts
    original_filename = file.filename
    # Sanitize filename (remove special characters except dots, hyphens, underscores)
    import re
    safe_filename = re.sub(r'[^\w\-_\.]', '_', original_filename)
    
    # Add unique prefix to prevent overwrites
    unique_prefix = str(uuid.uuid4())[:8]
    file_name = f"{unique_prefix}_{safe_filename}"
    file_path = UPLOAD_DIR / file_name
    
    with open(file_path, "wb") as buffer:
        shutil.copyfileobj(file.file, buffer)
    
    return {"file_id": file_name, "filename": original_filename, "original_name": original_filename}

@api_router.get("/uploads/{file_id}")
async def get_file(file_id: str):
    file_path = UPLOAD_DIR / file_id
    if not file_path.exists():
        raise HTTPException(status_code=404, detail="File not found")
    return FileResponse(file_path)

# ============ HELPER FUNCTION FOR DEPOT INVENTORY ============
# Note: This is used by the modular routes for liftings

async def update_depot_inventory(depot_id: str, depot_name: str, product_id: str, product_name: str, 
                                  product_code: str, quantity_change: float, is_incoming: bool):
    existing = await db.depot_inventory.find_one({
        "depot_id": depot_id,
        "product_id": product_id
    })
    
    if existing:
        if is_incoming:
            new_received = existing.get("total_received", 0) + quantity_change
            new_available = existing.get("available_quantity", 0) + quantity_change
            await db.depot_inventory.update_one(
                {"id": existing["id"]},
                {"$set": {
                    "total_received": new_received,
                    "available_quantity": new_available,
                    "last_updated": datetime.now(timezone.utc).isoformat()
                }}
            )
        else:
            new_dispatched = existing.get("total_dispatched", 0) + quantity_change
            new_available = existing.get("available_quantity", 0) - quantity_change
            await db.depot_inventory.update_one(
                {"id": existing["id"]},
                {"$set": {
                    "total_dispatched": new_dispatched,
                    "available_quantity": max(0, new_available),
                    "last_updated": datetime.now(timezone.utc).isoformat()
                }}
            )
    else:
        inventory = DepotInventory(
            depot_id=depot_id,
            depot_name=depot_name,
            product_id=product_id,
            product_name=product_name,
            product_code=product_code or "",
            total_received=quantity_change if is_incoming else 0,
            total_dispatched=0 if is_incoming else quantity_change,
            available_quantity=quantity_change if is_incoming else 0
        )
        await db.depot_inventory.insert_one(inventory.model_dump())

# ============ USERS ROUTES ============

@api_router.get("/users")
async def get_users(current_user: dict = Depends(get_current_user)):
    """Get all users - product and depot counts filtered by current user's access"""
    users = await db.users.find({}, {"_id": 0, "password": 0}).to_list(1000)

    # Get current user's accessible product and depot IDs
    admin_product_ids = await get_user_product_ids(current_user)
    admin_depot_ids = await get_user_depot_ids(current_user)

    products = await db.products.find({}, {"id": 1, "assigned_roles": 1}).to_list(1000)
    depots = await db.depots.find({}, {"id": 1, "assigned_roles": 1}).to_list(1000)

    admin_product_set = set(admin_product_ids) if admin_product_ids is not None else None
    admin_depot_set = set(admin_depot_ids) if admin_depot_ids is not None else None

    for user in users:
        role = user.get("role")
        user_id = user.get("id")

        effective_product_ids = set(user.get("assigned_products", []) or [])
        effective_depot_ids = set(user.get("assigned_depots", []) or [])

        if role:
            effective_product_ids.update([p["id"] for p in products if role in (p.get("assigned_roles") or [])])
            effective_depot_ids.update([d["id"] for d in depots if role in (d.get("assigned_roles") or [])])

        if admin_product_set is not None:
            effective_product_ids = {pid for pid in effective_product_ids if pid in admin_product_set}
        if admin_depot_set is not None:
            effective_depot_ids = {did for did in effective_depot_ids if did in admin_depot_set}

        user["effective_assigned_product_ids"] = list(effective_product_ids)
        user["effective_assigned_depot_ids"] = list(effective_depot_ids)

    return users

class UpdateUserRequest(BaseModel):
    name: Optional[str] = None
    role: Optional[str] = None
    email: Optional[str] = None
    depot_id: Optional[str] = None

@api_router.put("/users/{user_id}")
async def update_user(user_id: str, data: UpdateUserRequest, current_user: dict = Depends(get_current_user)):
    """Update user details (Management only)"""
    if current_user.get("role") != "Management":
        raise HTTPException(status_code=403, detail="Only Management can update users")
    
    user = await db.users.find_one({"id": user_id})
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    
    # Cannot modify master admin
    if user.get("is_master_admin"):
        raise HTTPException(status_code=403, detail="Cannot modify Master Admin")
    
    update_data = {}
    if data.name is not None:
        update_data["name"] = data.name
    if data.role is not None:
        update_data["role"] = data.role
    if data.email is not None:
        update_data["email"] = data.email
    if data.depot_id is not None:
        update_data["depot_id"] = data.depot_id
    
    if update_data:
        await db.users.update_one({"id": user_id}, {"$set": update_data})
    
    return {"success": True, "message": "User updated successfully"}

@api_router.delete("/users/{user_id}")
async def delete_user(user_id: str, current_user: dict = Depends(get_current_user)):
    """Delete user (Management only)"""
    if current_user.get("role") != "Management":
        raise HTTPException(status_code=403, detail="Only Management can delete users")
    
    user = await db.users.find_one({"id": user_id})
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    
    # Cannot delete master admin
    if user.get("is_master_admin"):
        raise HTTPException(status_code=403, detail="Cannot delete Master Admin")
    
    # Cannot delete self
    if user_id == current_user.get("id"):
        raise HTTPException(status_code=403, detail="Cannot delete your own account")
    
    await db.users.delete_one({"id": user_id})
    return {"message": "User deleted"}

# ============ EXPORT ROUTES ============

@api_router.get("/export/liftings")
async def export_liftings(
    format: str = "excel",
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    product_id: Optional[str] = None,
    delivery_order_id: Optional[str] = None,
    vehicle_id: Optional[str] = None,
    transporter_name: Optional[str] = None,
    loading_point_id: Optional[str] = None,
    unloading_point_id: Optional[str] = None,
    unloading_status: Optional[str] = None,
    lifting_type: Optional[str] = None
):
    """Export liftings data to Excel or PDF"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    
    # Build query
    query = {}
    if date_from or date_to:
        date_query = {}
        if date_from:
            date_query["$gte"] = date_from
        if date_to:
            date_query["$lte"] = date_to
        if date_query:
            query["date_of_loading"] = date_query
    if product_id:
        query["product_id"] = product_id
    if delivery_order_id:
        query["delivery_order_id"] = delivery_order_id
    if vehicle_id:
        query["vehicle_id"] = vehicle_id
    if transporter_name:
        query["transporter_name"] = {"$regex": transporter_name, "$options": "i"}
    if loading_point_id:
        query["loading_point_id"] = loading_point_id
    if unloading_point_id:
        query["unloading_point_id"] = unloading_point_id
    if unloading_status:
        query["unloading_status"] = unloading_status
    if lifting_type:
        query["lifting_type"] = lifting_type
    
    liftings = await db.liftings.find(query, {"_id": 0}).to_list(10000)
    
    if format == "excel":
        wb = Workbook()
        ws = wb.active
        ws.title = "Liftings Report"
        
        # Header styling
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Headers
        headers = ["Lifting No", "Date", "Type", "Product", "Quantity (MT)", "Vehicle", 
                   "Transporter", "Driver", "From", "To", "Status", "Net Weight (MT)"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border
        
        # Data rows
        for row, lifting in enumerate(liftings, 2):
            ws.cell(row=row, column=1, value=lifting.get("lifting_no", ""))
            ws.cell(row=row, column=2, value=lifting.get("date_of_loading", ""))
            ws.cell(row=row, column=3, value=lifting.get("lifting_type", ""))
            ws.cell(row=row, column=4, value=lifting.get("product_name", ""))
            ws.cell(row=row, column=5, value=lifting.get("quantity_mt", 0))
            ws.cell(row=row, column=6, value=lifting.get("vehicle_number", ""))
            ws.cell(row=row, column=7, value=lifting.get("transporter_name", ""))
            ws.cell(row=row, column=8, value=lifting.get("driver_name", ""))
            ws.cell(row=row, column=9, value=lifting.get("loading_point_name", ""))
            ws.cell(row=row, column=10, value=lifting.get("unloading_point_name", ""))
            ws.cell(row=row, column=11, value=lifting.get("unloading_status", ""))
            ws.cell(row=row, column=12, value=lifting.get("net_weight_mt", 0))
            for col in range(1, 13):
                ws.cell(row=row, column=col).border = thin_border
        
        # Auto-width columns
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 30)
            ws.column_dimensions[column].width = adjusted_width
        
        # Save to buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        filename = f"liftings_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        return StreamingResponse(
            buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    
    return {"error": "Unsupported format"}

@api_router.get("/export/inventory")
async def export_inventory(format: str = "excel"):
    """Export depot inventory to Excel"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    
    inventory = await db.depot_inventory.find({}, {"_id": 0}).to_list(10000)
    
    if format == "excel":
        wb = Workbook()
        ws = wb.active
        ws.title = "Inventory Report"
        
        # Header styling
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="059669", end_color="059669", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Headers
        headers = ["Depot Name", "Product Name", "Quantity (MT)", "Last Updated"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border
        
        # Data rows
        for row, item in enumerate(inventory, 2):
            ws.cell(row=row, column=1, value=item.get("depot_name", ""))
            ws.cell(row=row, column=2, value=item.get("product_name", ""))
            ws.cell(row=row, column=3, value=item.get("quantity_mt", 0))
            ws.cell(row=row, column=4, value=item.get("updated_at", ""))
            for col in range(1, 5):
                ws.cell(row=row, column=col).border = thin_border
        
        # Auto-width columns
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 30)
            ws.column_dimensions[column].width = adjusted_width
        
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        filename = f"inventory_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        return StreamingResponse(
            buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    
    return {"error": "Unsupported format"}

@api_router.get("/export/delivery-orders")
async def export_delivery_orders(format: str = "excel", status: Optional[str] = None):
    """Export delivery orders summary to Excel"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    
    query = {}
    if status:
        query["status"] = status
    if not current_user.get("is_master_admin"):
        query["from_company_id"] = current_user["company_id"]

    
    orders = await db.delivery_orders.find(query, {"_id": 0}).to_list(10000)
    
    if format == "excel":
        wb = Workbook()
        ws = wb.active
        ws.title = "Delivery Orders"
        
        # Header styling
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="7C3AED", end_color="7C3AED", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Headers
        headers = ["DO Number", "Date", "Product", "Company", "Total Qty (MT)", 
                   "Remaining Qty (MT)", "To Depot", "Status"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border
        
        # Data rows
        for row, order in enumerate(orders, 2):
            ws.cell(row=row, column=1, value=order.get("do_order_no", ""))
            ws.cell(row=row, column=2, value=order.get("do_date", ""))
            ws.cell(row=row, column=3, value=order.get("product_name", ""))
            ws.cell(row=row, column=4, value=order.get("from_company_name", ""))
            ws.cell(row=row, column=5, value=order.get("total_quantity_mt", 0))
            ws.cell(row=row, column=6, value=order.get("remaining_quantity_mt", 0))
            ws.cell(row=row, column=7, value=order.get("to_depot_name", ""))
            ws.cell(row=row, column=8, value=order.get("status", ""))
            for col in range(1, 9):
                ws.cell(row=row, column=col).border = thin_border
        
        # Auto-width columns
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 30)
            ws.column_dimensions[column].width = adjusted_width
        
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        filename = f"delivery_orders_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        return StreamingResponse(
            buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    
    return {"error": "Unsupported format"}

@api_router.get("/export/users")
async def export_users(format: str = "excel"):
    """Export all users data to Excel"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    
    users = await db.users.find({}, {"_id": 0, "password": 0}).to_list(10000)
    
    if format == "excel":
        wb = Workbook()
        ws = wb.active
        ws.title = "Users Report"
        
        # Header styling
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Headers (removed Assigned Products and Is Master Admin)
        headers = [
            "S.No", "Name", "Mobile", "Country Code", "Email", "Role", 
            "Depot ID", "OTP Verified", "Password Set", "Created At"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # Data rows
        for row_idx, user in enumerate(users, 2):
            row_data = [
                row_idx - 1,  # S.No
                user.get('name', ''),
                user.get('mobile', ''),
                user.get('country_code', '91'),
                user.get('email', ''),
                user.get('role', ''),
                user.get('depot_id', ''),
                'Yes' if user.get('otp_verified') else 'No',
                'Yes' if user.get('password_set') else 'No',
                user.get('created_at', '')[:19] if user.get('created_at') else ''
            ]
            
            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col, value=value)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="left", vertical="center")
        
        # Adjust column widths
        column_widths = [8, 25, 15, 12, 30, 15, 20, 12, 12, 20]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[chr(64 + col) if col <= 26 else 'A' + chr(64 + col - 26)].width = width
        
        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = f"users_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    
    return {"error": "Unsupported format"}

@api_router.get("/purchase-orders/{order_id}/statement/export")
async def export_purchase_order_statement(
    order_id: str,
    format: str = "excel"
):
    """Export purchase order statement to Excel or PDF"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    
    # TODO: Re-enable authentication and permission checks after testing
    # await check_permission(current_user, "Purchase Orders (View)")
    
    # Get PO
    order = await db.purchase_orders.find_one(
        {"id": order_id},
        {"_id": 0}
    )
    
    if not order:
        raise HTTPException(status_code=404, detail="Purchase Order not found")
    
    # Product access check - TODO: Re-enable after testing
    # if order.get("product_id"):
    #     await check_product_access(current_user, order["product_id"])
    
    # Get liftings
    liftings = await db.liftings.find(
        {
            "purchase_order_id": order_id,
            "unloading_status": {"$ne": "Rejected"}
        },
        {"_id": 0}
    ).sort("date_of_loading", -1).to_list(1000)
    
    # Get verified pickups
    pickups = await db.pickups.find(
        {
            "purchase_order_id": order_id,
            "status": "verified"
        },
        {"_id": 0}
    ).sort("verified_at", -1).to_list(1000)
    
    transactions = []
    
    # Add pickup rows
    for p in pickups:
        transactions.append({
            "date": p.get("verified_at"),
            "type": "Pickup",
            "reference_no": p.get("purchase_order_no"),
            "vehicle": p.get("truck_number"),
            "quantity": p.get("weight_mt", 0),
            "status": p.get("status", "")
        })
    
    # Add lifting rows
    for l in liftings:
        transactions.append({
            "date": l.get("date_of_loading"),
            "type": "Lifting",
            "reference_no": l.get("lifting_no"),
            "vehicle": l.get("vehicle_id"),
            "quantity": l.get("net_weight_mt", 0),
            "status": l.get("unloading_status", "")
        })
    
    # Sort by date
    transactions.sort(key=lambda x: x.get("date", ""), reverse=True)
    
    if format == "excel":
        wb = Workbook()
        ws = wb.active
        ws.title = "PO Statement"
        
        # Header styling
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
        title_font = Font(bold=True, size=14)
        summary_font = Font(bold=True, size=10)
        summary_fill = PatternFill(start_color="E0E7FF", end_color="E0E7FF", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Title
        ws.merge_cells("A1:F1")
        title_cell = ws["A1"]
        title_cell.value = f"Purchase Order Statement - {order.get('po_number', 'N/A')}"
        title_cell.font = title_font
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Summary section
        row = 3
        ws[f"A{row}"] = "PO Number:"
        ws[f"B{row}"] = order.get("po_number", "")
        row += 1
        ws[f"A{row}"] = "Total PO Quantity (MT):"
        ws[f"B{row}"] = order.get("total_quantity_mt", 0)
        row += 1
        ws[f"A{row}"] = "Lifted (MT):"
        ws[f"B{row}"] = sum(t.get("quantity", 0) for t in transactions if t.get("type") == "Lifting")
        row += 1
        ws[f"A{row}"] = "Picked Up (MT):"
        ws[f"B{row}"] = sum(t.get("quantity", 0) for t in transactions if t.get("type") == "Pickup")
        row += 1
        ws[f"A{row}"] = "Status:"
        ws[f"B{row}"] = order.get("status", "")
        
        # Transactions table header
        row = 10
        headers = ["Date", "Type", "Reference", "Vehicle", "Quantity (MT)", "Status"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border
        
        # Data rows
        row = 11
        for tx in transactions:
            ws.cell(row=row, column=1, value=tx.get("date", "")[:10] if tx.get("date") else "")
            ws.cell(row=row, column=2, value=tx.get("type", ""))
            ws.cell(row=row, column=3, value=tx.get("reference_no", ""))
            ws.cell(row=row, column=4, value=tx.get("vehicle", ""))
            ws.cell(row=row, column=5, value=tx.get("quantity", 0))
            ws.cell(row=row, column=6, value=tx.get("status", ""))
            
            for col in range(1, 7):
                cell = ws.cell(row=row, column=col)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='left', vertical='center')
            
            row += 1
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 15
        
        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = f"PO_Statement_{order.get('po_number', 'N/A')}.xlsx"
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    
    elif format == "pdf":
        try:
            import datetime
            from reportlab.lib.pagesizes import A4
            from reportlab.lib import colors
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import inch
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
            
            output = io.BytesIO()
            # Set precise, uniform margins for full presentation control
            doc = SimpleDocTemplate(output, pagesize=A4, leftMargin=0.5*inch, rightMargin=0.5*inch, topMargin=0.5*inch, bottomMargin=0.5*inch)
            story = []
            
            styles = getSampleStyleSheet()
            
            # Professional corporate typography layout
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=18,
                textColor=colors.black,
                spaceAfter=4,
                fontName='Helvetica-Bold',
                alignment=TA_LEFT
            )
            
            meta_left_style = ParagraphStyle(
                'MetaLeft',
                parent=styles['Normal'],
                fontSize=10,
                textColor=colors.HexColor('#DC2626'), # Red color matching your boss's notes
                fontName='Helvetica-Bold',
                leading=14
            )
            
            meta_right_style = ParagraphStyle(
                'MetaRight',
                parent=styles['Normal'],
                fontSize=10,
                textColor=colors.HexColor('#DC2626'), # Red color matching your boss's notes
                fontName='Helvetica-Bold',
                alignment=TA_RIGHT,
                leading=14
            )

            table_header_style = ParagraphStyle(
                'TableHeader',
                parent=styles['Normal'],
                fontSize=9,
                textColor=colors.whitesmoke,
                fontName='Helvetica-Bold',
                alignment=TA_CENTER
            )

            table_cell_style = ParagraphStyle(
                'TableCell',
                parent=styles['Normal'],
                fontSize=9,
                textColor=colors.HexColor('#1E293B'),
                fontName='Helvetica',
                alignment=TA_CENTER
            )
            
            table_cell_right = ParagraphStyle(
                'TableCellRight',
                parent=table_cell_style,
                alignment=TA_RIGHT,
                fontName='Helvetica-Bold'
            )

            # 1. Main Document Title Row
            title = Paragraph(f"Purchase Order Statement - {order.get('po_number', 'N/A')}", title_style)
            story.append(title)
            story.append(Spacer(1, 0.1*inch))
            
            # 2. Header Metadata (Company, Product vs Current Download Date)
            current_date_str = datetime.datetime.now().strftime("%d/%m/%Y")
            
            meta_data = [
                [
                    Paragraph(f"Company: {order.get('to_company_name', 'N/A')}<br/>Product: {order.get('product_name', 'N/A')}", meta_left_style),
                    Paragraph(f"Date: {current_date_str}", meta_right_style)
                ]
            ]
            
            # 7.25 inches is the precise maximum width on A4 page with 0.5-inch margins
            meta_table = Table(meta_data, colWidths=[4.25*inch, 3.0*inch])
            meta_table.setStyle(TableStyle([
                ('VALIGN', (0,0), (-1,-1), 'TOP'),
                ('LEFTPADDING', (0,0), (-1,-1), 0),
                ('RIGHTPADDING', (0,0), (-1,-1), 0),
                ('BOTTOMPADDING', (0,0), (-1,-1), 0),
                ('TOPPADDING', (0,0), (-1,-1), 0),
            ]))
            story.append(meta_table)
            story.append(Spacer(1, 0.2*inch))
            
            # 3. Dynamic Summary Performance Table Card
            total_dispatch_qty = sum(t.get("quantity", 0) for t in transactions)
            raw_po_date = order.get("po_date", "")
            formatted_po_date = raw_po_date[:10] if raw_po_date else "N/A"
            
            summary_data = [
                ["PO Number", order.get("po_number", "N/A")],
                ["Total PO Quantity (MT)", f"{order.get('total_quantity_mt', 0)} MT"],
                ["PO Date", formatted_po_date],
                ["Dispatch Quantity (MT)", f"{total_dispatch_qty:.2f} MT"],
                ["Status", str(order.get("status", "N/A")).upper()]
            ]
            
            summary_table = Table(summary_data, colWidths=[2.5*inch, 4.75*inch])
            summary_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#E0E7FF')), # Light indigo/blue label fill
                ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 9.5),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('TOPPADDING', (0, 0), (-1, -1), 6),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
                ('LEFTPADDING', (0, 0), (-1, -1), 10),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#CBD5E1')) # Clean modern border color instead of pure harsh black
            ]))
            
            story.append(summary_table)
            story.append(Spacer(1, 0.25*inch))
            
            # 4. Itemized Transaction Ledger Grid (Your colors, but beautiful spacing)
            table_data = [[
                Paragraph("Date", table_header_style), 
                Paragraph("Type", table_header_style), 
                Paragraph("Reference", table_header_style), 
                Paragraph("Vehicle", table_header_style), 
                Paragraph("Quantity (MT)", table_header_style), 
                Paragraph("Status", table_header_style)
            ]]
            
            for tx in transactions:
                raw_type = tx.get("type", "")
                display_type = "Dispatch" if raw_type == "Pickup" else raw_type
                
                table_data.append([
                    Paragraph(tx.get("date", "")[:10] if tx.get("date") else "-", table_cell_style),
                    Paragraph(display_type, table_cell_style),
                    Paragraph(tx.get("reference_no", "-"), table_cell_style),
                    Paragraph(tx.get("vehicle", "-"), table_cell_style),
                    Paragraph(f"{float(tx.get('quantity', 0)):.2f}", table_cell_right),
                    Paragraph(str(tx.get("status", "-")).capitalize(), table_cell_style)
                ])
            
            # Perfect A4 layout width scaling
            table = Table(table_data, colWidths=[1.1*inch, 1.1*inch, 1.3*inch, 1.1*inch, 1.3*inch, 1.35*inch])
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2563EB')), # Your exact vibrant blue header
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('TOPPADDING', (0, 0), (-1, -1), 8),    # Better inner padding
                ('BOTTOMPADDING', (0, 0), (-1, -1), 8), # Better inner padding
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige), # Your clean exact beige backdrop background choice
                ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#94A3B8')), # Professional soft grid lines
            ]))
            
            story.append(table)
            doc.build(story)
            output.seek(0)
            
            filename = f"PO_Statement_{order.get('po_number', 'N/A')}.pdf"
            return StreamingResponse(
                output,
                media_type="application/pdf",
                headers={"Content-Disposition": f"attachment; filename={filename}"}
            )
        except ImportError:
            return {"error": "PDF export requires reportlab library"}
    
    return {"error": "Unsupported format"}

# ============ BULK IMPORT ROUTES ============

@api_router.get("/import/template/{entity}")
async def get_import_template(entity: str):
    """Get Excel template for bulk import"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    
    wb = Workbook()
    ws = wb.active
    
    # Header styling
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    templates = {
        "trucks": {
            "title": "Trucks Import Template",
            "headers": ["Vehicle Number*", "Transporter Name", "Driver Name", "Driver Mobile", 
                       "Helper Name", "Helper Mobile", "Tare Weight (MT)", "Capacity (MT)"],
            "example": ["MH12AB1234", "ABC Transport", "John Doe", "9876543210", 
                       "Helper Name", "9876543211", "5.5", "10"]
        },
        "products": {
            "title": "Products Import Template",
            "headers": ["Product Name*", "Product Code*", "Category", "HSN Code", "Unit", "Description"],
            "example": ["Cement", "CEM001", "Building Materials", "2523", "MT", "Portland Cement"]
        },
        "companies": {
            "title": "Companies Import Template",
            "headers": ["Company Name*", "Address", "City", "State", "Country", "PIN Code", 
                       "Contact Person", "Phone", "Email"],
            "example": ["ABC Corp", "123 Main St", "Mumbai", "Maharashtra", "India", "400001", 
                       "John Doe", "9876543210", "contact@abc.com"]
        },
        "transporters": {
            "title": "Transporters Import Template",
            "headers": ["Transporter Name*", "Contact Person", "Phone", "Email", "Address", "GST Number"],
            "example": ["XYZ Transport", "Manager Name", "9876543210", "xyz@transport.com", 
                       "456 Transport Hub", "27XXXXX1234X1Z5"]
        }
    }
    
    if entity not in templates:
        raise HTTPException(status_code=400, detail=f"Unknown entity: {entity}")
    
    template = templates[entity]
    ws.title = template["title"]
    
    # Add headers
    for col, header in enumerate(template["headers"], 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
    
    # Add example row
    for col, value in enumerate(template["example"], 1):
        cell = ws.cell(row=2, column=col, value=value)
        cell.border = thin_border
    
    # Add instructions
    ws.cell(row=4, column=1, value="Instructions:")
    ws.cell(row=5, column=1, value="* = Required field")
    ws.cell(row=6, column=1, value="Delete the example row before importing")
    ws.cell(row=7, column=1, value="Do not modify headers")
    
    # Auto-width columns
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 30)
        ws.column_dimensions[column].width = adjusted_width
    
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    filename = f"{entity}_import_template.xlsx"
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

@api_router.post("/import/{entity}")
async def bulk_import(entity: str, file: UploadFile = File(...)):
    """Bulk import data from Excel"""
    from openpyxl import load_workbook
    
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Only Excel files (.xlsx, .xls) are supported")
    
    try:
        contents = await file.read()
        wb = load_workbook(io.BytesIO(contents))
        ws = wb.active
        
        # Get headers from first row
        headers = [cell.value for cell in ws[1]]
        
        imported = 0
        errors = []
        
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
            # Skip empty rows
            if not any(row):
                continue
                
            row_data = dict(zip(headers, row))
            
            try:
                if entity == "trucks":
                    vehicle_number = row_data.get("Vehicle Number*")
                    if not vehicle_number:
                        errors.append(f"Row {row_idx}: Vehicle Number is required")
                        continue
                    
                    # Check if truck already exists
                    existing = await db.trucks.find_one({"vehicle_number": vehicle_number})
                    if existing:
                        errors.append(f"Row {row_idx}: Vehicle {vehicle_number} already exists")
                        continue
                    
                    truck = {
                        "id": str(uuid.uuid4()),
                        "vehicle_number": vehicle_number,
                        "transporter_name": row_data.get("Transporter Name") or "",
                        "driver_name": row_data.get("Driver Name") or "",
                        "driver_mobile": row_data.get("Driver Mobile") or "",
                        "helper_name": row_data.get("Helper Name") or "",
                        "helper_mobile": row_data.get("Helper Mobile") or "",
                        "tare_weight_mt": float(row_data.get("Tare Weight (MT)") or 0),
                        "capacity_mt": float(row_data.get("Capacity (MT)") or 0),
                        "created_at": datetime.now(timezone.utc).isoformat()
                    }
                    await db.trucks.insert_one(truck)
                    imported += 1
                    
                elif entity == "products":
                    product_name = row_data.get("Product Name*")
                    product_code = row_data.get("Product Code*")
                    if not product_name or not product_code:
                        errors.append(f"Row {row_idx}: Product Name and Code are required")
                        continue
                    
                    # Check if product code already exists
                    existing = await db.products.find_one({"product_code": product_code})
                    if existing:
                        errors.append(f"Row {row_idx}: Product code {product_code} already exists")
                        continue
                    
                    product = {
                        "id": str(uuid.uuid4()),
                        "product_name": product_name,
                        "product_code": product_code,
                        "category": row_data.get("Category") or "",
                        "hsn_code": row_data.get("HSN Code") or "",
                        "unit": row_data.get("Unit") or "MT",
                        "description": row_data.get("Description") or "",
                        "created_at": datetime.now(timezone.utc).isoformat()
                    }
                    await db.products.insert_one(product)
                    imported += 1
                    
                elif entity == "companies":
                    company_name = row_data.get("Company Name*")
                    if not company_name:
                        errors.append(f"Row {row_idx}: Company Name is required")
                        continue
                    
                    company = {
                        "id": str(uuid.uuid4()),
                        "name": company_name,
                        "address": row_data.get("Address") or "",
                        "city": row_data.get("City") or "",
                        "state": row_data.get("State") or "",
                        "country": row_data.get("Country") or "India",
                        "pin_code": row_data.get("PIN Code") or "",
                        "contact_person": row_data.get("Contact Person") or "",
                        "phone": row_data.get("Phone") or "",
                        "email": row_data.get("Email") or "",
                        "created_at": datetime.now(timezone.utc).isoformat()
                    }
                    await db.companies.insert_one(company)
                    imported += 1
                    
                elif entity == "transporters":
                    transporter_name = row_data.get("Transporter Name*")
                    if not transporter_name:
                        errors.append(f"Row {row_idx}: Transporter Name is required")
                        continue
                    
                    transporter = {
                        "id": str(uuid.uuid4()),
                        "name": transporter_name,
                        "contact_person": row_data.get("Contact Person") or "",
                        "phone": row_data.get("Phone") or "",
                        "email": row_data.get("Email") or "",
                        "address": row_data.get("Address") or "",
                        "gst_number": row_data.get("GST Number") or "",
                        "created_at": datetime.now(timezone.utc).isoformat()
                    }
                    await db.transporters.insert_one(transporter)
                    imported += 1
                    
                else:
                    raise HTTPException(status_code=400, detail=f"Unknown entity: {entity}")
                    
            except Exception as e:
                errors.append(f"Row {row_idx}: {str(e)}")
        
        return {
            "success": True,
            "imported": imported,
            "errors": errors[:10],  # Return first 10 errors
            "total_errors": len(errors)
        }
        
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Failed to process file: {str(e)}")

# ============ ANALYTICS ROUTES ============

@api_router.get("/analytics/dashboard")
async def get_dashboard_analytics(current_user: dict = Depends(get_current_user)):
    companies_count = await db.companies.count_documents({})
    users_count = await db.users.count_documents({})
    transporters_count = await db.transporters.count_documents({})
    trucks_count = await db.trucks.count_documents({})
    products_count = await db.products.count_documents({})
    depots_count = await db.depots.count_documents({})
    orders_count = await db.delivery_orders.count_documents({})
    liftings_count = await db.liftings.count_documents({})
    query = {}
    if not current_user.get("is_master_admin"):
        query["from_company_id"] = current_user["company_id"]
    query["status"] = "Open"    
    open_orders = await db.delivery_orders.count_documents(query)
    query["status"] = "In Progress"
    in_progress = await db.delivery_orders.count_documents(query)
    query["status"] = "Completed"
    completed = await db.delivery_orders.count_documents(query)
    
    pending_verification = await db.liftings.count_documents({"unloading_status": "Pending"})
    verified = await db.liftings.count_documents({"unloading_status": "Verified"})

    # Product access filter for product-level dashboard metrics
    product_filter = await build_product_filter(current_user, "product_id")

    # Get detailed delivery orders breakdown (Open status with product info)
    open_do_query = {"status": {"$in": ["Open", "In Progress"]}}
    open_do_query.update(product_filter)
    open_dos = await db.delivery_orders.find(
        open_do_query,
        {"_id": 0, "id": 1, "do_order_no": 1, "product_name": 1, "product_id": 1, 
         "total_quantity_mt": 1, "lifted_quantity_mt": 1, "remaining_quantity_mt": 1, "status": 1}
    ).to_list(100)
    
    # Use aggregation for efficient DO quantity calculation by product
    do_qty_pipeline = [
        {"$match": product_filter},
        {"$group": {
            "_id": "$product_id",
            "product_name": {"$first": "$product_name"},
            "total_qty": {"$sum": "$total_quantity_mt"}
        }}
    ]
    do_qty_result = await db.delivery_orders.aggregate(do_qty_pipeline).to_list(100)
    do_qty_by_product = {
        item["_id"] or "unknown": {"product_name": item.get("product_name") or "Unknown", "total_qty": item.get("total_qty") or 0}
        for item in do_qty_result
    }

    # Remaining DO quantity by product
    do_remaining_pipeline = [
        {"$match": {**product_filter, "status": {"$ne": "Completed"}}},
        {"$group": {
            "_id": "$product_id",
            "product_name": {"$first": "$product_name"},
            "remaining_qty": {"$sum": "$remaining_quantity_mt"}
        }}
    ]
    do_remaining_result = await db.delivery_orders.aggregate(do_remaining_pipeline).to_list(200)
    do_remaining_by_product = {
        item["_id"] or "unknown": {"product_name": item.get("product_name") or "Unknown", "remaining_qty": item.get("remaining_qty") or 0}
        for item in do_remaining_result
    }

    # Available stock by product from depot inventory
    inventory_pipeline = [
        {"$match": product_filter},
        {"$group": {
            "_id": "$product_id",
            "product_name": {"$first": "$product_name"},
            "available_stock": {"$sum": "$available_quantity"}
        }}
    ]
    inventory_result = await db.depot_inventory.aggregate(inventory_pipeline).to_list(200)
    available_stock_by_product = {
        item["_id"] or "unknown": {"product_name": item.get("product_name") or "Unknown", "available_stock": item.get("available_stock") or 0}
        for item in inventory_result
    }

    depot_dispatch_by_product = {}

    # Remaining PO quantity by product
    po_remaining_pipeline = [
        {"$match": {**product_filter, "status": {"$ne": "Completed"}}},
        {"$group": {
            "_id": "$product_id",
            "product_name": {"$first": "$product_name"},
            "remaining_qty": {"$sum": "$remaining_quantity_mt"}
        }}
    ]
    po_remaining_result = await db.purchase_orders.aggregate(po_remaining_pipeline).to_list(200)
    po_remaining_by_product = {
        item["_id"] or "unknown": {"product_name": item.get("product_name") or "Unknown", "remaining_qty": item.get("remaining_qty") or 0}
        for item in po_remaining_result
    }

    # Dispatch quantity for today, yesterday, and day before yesterday
    today = datetime.now(timezone.utc)
    today_key = today.strftime("%Y-%m-%d")
    yesterday = today - timedelta(days=1)
    yesterday_key = yesterday.strftime("%Y-%m-%d")
    day_before = today - timedelta(days=2)
    day_before_key = day_before.strftime("%Y-%m-%d")
    dispatch_start = day_before.strftime("%Y-%m-%d") + "T00:00:00"
    dispatch_end = today.strftime("%Y-%m-%d") + "T23:59:59"

    dispatch_by_product = {}

    secondary_dispatch_query = {
        **product_filter,
        "lifting_type": "Secondary",
        "date_of_loading": {"$gte": dispatch_start, "$lte": dispatch_end}
    }
    secondary_dispatches = await db.liftings.find(
        secondary_dispatch_query,
        {"_id": 0, "product_id": 1, "product_name": 1, "quantity_mt": 1, "date_of_loading": 1, "loading_point_id": 1, "loading_point_name": 1}
    ).to_list(1000)

    pickup_dispatch_query = {
        **product_filter,
        "status": "verified",
        "verified_at": {"$gte": dispatch_start, "$lte": dispatch_end}
    }
    # include depot info so pickups contribute to depot-level dispatch totals
    pickup_dispatches = await db.pickups.find(
        pickup_dispatch_query,
        {"_id": 0, "product_id": 1, "product_name": 1, "weight_mt": 1, "verified_at": 1, "depot_id": 1, "depot_name": 1}
    ).to_list(1000)

    def add_dispatch(product_id, name, qty, date_key):
        pid = product_id or "unknown"
        if pid not in dispatch_by_product:
            dispatch_by_product[pid] = {
                "product_name": name or "Unknown",
                "today": 0,
                "yesterday": 0,
                "day_before_yesterday": 0
            }
        dispatch_by_product[pid][date_key] += qty

    depot_dispatch_by_product = {}

    def add_depot_dispatch(product_id, depot_id, depot_name, qty, date_key):
        pid = product_id or "unknown"
        if pid not in depot_dispatch_by_product:
            depot_dispatch_by_product[pid] = {}
        did = depot_id or "unknown"
        if did not in depot_dispatch_by_product[pid]:
            depot_dispatch_by_product[pid][did] = {
                "depot_id": did,
                "depot_name": depot_name or "Unknown Depot",
                "dispatch_today_qty": 0,
                "dispatch_yesterday_qty": 0,
                "dispatch_day_before_yesterday_qty": 0
            }
        depot_dispatch_by_product[pid][did][date_key] += qty

    for item in secondary_dispatches:
        date_value = (item.get("date_of_loading") or "")[:10]
        qty = item.get("quantity_mt") or 0
        if date_value == today_key:
            add_dispatch(item.get("product_id"), item.get("product_name"), qty, "today")
            add_depot_dispatch(item.get("product_id"), item.get("loading_point_id"), item.get("loading_point_name"), qty, "dispatch_today_qty")
        elif date_value == yesterday_key:
            add_dispatch(item.get("product_id"), item.get("product_name"), qty, "yesterday")
            add_depot_dispatch(item.get("product_id"), item.get("loading_point_id"), item.get("loading_point_name"), qty, "dispatch_yesterday_qty")
        elif date_value == day_before_key:
            add_dispatch(item.get("product_id"), item.get("product_name"), qty, "day_before_yesterday")
            add_depot_dispatch(item.get("product_id"), item.get("loading_point_id"), item.get("loading_point_name"), qty, "dispatch_day_before_yesterday_qty")

    for item in pickup_dispatches:
        date_value = (item.get("verified_at") or "")[:10]
        qty = item.get("weight_mt") or 0
        if date_value == today_key:
            add_dispatch(item.get("product_id"), item.get("product_name"), qty, "today")
            # attribute pickup dispatch to depot if present
            if item.get("depot_id"):
                add_depot_dispatch(item.get("product_id"), item.get("depot_id"), item.get("depot_name"), qty, "dispatch_today_qty")
        elif date_value == yesterday_key:
            add_dispatch(item.get("product_id"), item.get("product_name"), qty, "yesterday")
            if item.get("depot_id"):
                add_depot_dispatch(item.get("product_id"), item.get("depot_id"), item.get("depot_name"), qty, "dispatch_yesterday_qty")
        elif date_value == day_before_key:
            add_dispatch(item.get("product_id"), item.get("product_name"), qty, "day_before_yesterday")
            if item.get("depot_id"):
                add_depot_dispatch(item.get("product_id"), item.get("depot_id"), item.get("depot_name"), qty, "dispatch_day_before_yesterday_qty")

    inventory_details = await db.depot_inventory.find(
        product_filter,
        {"_id": 0, "product_id": 1, "depot_id": 1, "depot_name": 1, "available_quantity": 1}
    ).to_list(1000)
    stock_by_depot_by_product = {}
    for item in inventory_details:
        pid = item.get("product_id") or "unknown"
        depot_id = item.get("depot_id") or "unknown"
        depot_dispatch = depot_dispatch_by_product.get(pid, {}).get(depot_id, {})
        stock_by_depot_by_product.setdefault(pid, []).append({
            "depot_id": depot_id,
            "depot_name": item.get("depot_name") or "Unknown Depot",
            "available_quantity": item.get("available_quantity") or 0,
            "dispatch_today_qty": depot_dispatch.get("dispatch_today_qty", 0),
            "dispatch_yesterday_qty": depot_dispatch.get("dispatch_yesterday_qty", 0),
            "dispatch_day_before_yesterday_qty": depot_dispatch.get("dispatch_day_before_yesterday_qty", 0)
        })

    # Use aggregation for efficient Primary liftings breakdown
    primary_pipeline = [
        {"$match": {"lifting_type": "Primary"}},
        {"$group": {
            "_id": {"product_id": "$product_id", "unloading_point_type": "$unloading_point_type"},
            "product_name": {"$first": "$product_name"},
            "total_qty": {"$sum": "$quantity_mt"},
            "count": {"$sum": 1}
        }}
    ]
    primary_result = await db.liftings.aggregate(primary_pipeline).to_list(200)
    
    # Use aggregation for efficient Secondary liftings breakdown
    secondary_pipeline = [
        {"$match": {"lifting_type": "Secondary"}},
        {"$group": {
            "_id": {"product_id": "$product_id", "unloading_point_type": "$unloading_point_type"},
            "product_name": {"$first": "$product_name"},
            "total_qty": {"$sum": "$quantity_mt"},
            "count": {"$sum": 1}
        }}
    ]
    secondary_result = await db.liftings.aggregate(secondary_pipeline).to_list(200)
    
    # Build primary liftings breakdown from aggregation results
    primary_liftings = []
    primary_to_depot_count = 0
    primary_to_client_count = 0
    for item in primary_result:
        qty = item.get("total_qty") or 0
        count = item.get("count") or 0
        unloading_type = item["_id"].get("unloading_point_type")
        if unloading_type == "Depot":
            primary_to_depot_count += count
        else:
            primary_to_client_count += count
        # Add to list for compatibility with existing code
        for _ in range(count):
            primary_liftings.append({
                "product_id": item["_id"].get("product_id"),
                "product_name": item.get("product_name"),
                "quantity_mt": qty / count if count > 0 else 0,
                "unloading_point_type": unloading_type
            })
    
    # Build secondary liftings breakdown from aggregation results
    secondary_liftings = []
    for item in secondary_result:
        qty = item.get("total_qty") or 0
        count = item.get("count") or 0
        for _ in range(count):
            secondary_liftings.append({
                "product_id": item["_id"].get("product_id"),
                "product_name": item.get("product_name"),
                "quantity_mt": qty / count if count > 0 else 0,
                "unloading_point_type": item["_id"].get("unloading_point_type")
            })
    
    # Primary liftings breakdown: To Depot vs To Client (Direct)
    primary_to_depot = [l for l in primary_liftings if l.get("unloading_point_type") == "Depot"]
    primary_to_client = [l for l in primary_liftings if l.get("unloading_point_type") != "Depot"]
    
    # Calculate lifted qty per product for Primary
    primary_by_product = {}
    for l in primary_liftings:
        pid = l.get("product_id") or "unknown"
        pname = l.get("product_name") or "Unknown"
        qty = l.get("quantity_mt") or 0
        if pid not in primary_by_product:
            primary_by_product[pid] = {"product_name": pname, "lifted_qty": 0, "to_depot": 0, "to_client": 0}
        primary_by_product[pid]["lifted_qty"] += qty
        if l.get("unloading_point_type") == "Depot":
            primary_by_product[pid]["to_depot"] += qty
        else:
            primary_by_product[pid]["to_client"] += qty
    
    # Calculate lifted qty per product for Secondary
    secondary_by_product = {}
    for l in secondary_liftings:
        pid = l.get("product_id") or "unknown"
        pname = l.get("product_name") or "Unknown"
        qty = l.get("quantity_mt") or 0
        if pid not in secondary_by_product:
            secondary_by_product[pid] = {"product_name": pname, "lifted_qty": 0, "to_company": 0, "to_depot": 0}
        secondary_by_product[pid]["lifted_qty"] += qty
        if l.get("unloading_point_type") == "Company":
            secondary_by_product[pid]["to_company"] += qty
        elif l.get("unloading_point_type") == "Depot":
            secondary_by_product[pid]["to_depot"] += qty
    
    # Secondary liftings breakdown: To Company vs To Depot
    secondary_to_company = [l for l in secondary_liftings if l.get("unloading_point_type") == "Company"]
    secondary_to_depot = [l for l in secondary_liftings if l.get("unloading_point_type") == "Depot"]
    
    # Use aggregation for company deliveries (optimized query with limit)
    company_deliveries_pipeline = [
        {"$match": {"unloading_point_type": "Company"}},
        {"$sort": {"date_of_loading": -1}},
        {"$limit": 500},  # Limit to recent 500 deliveries for dashboard
        {"$project": {
            "_id": 0, "product_id": 1, "product_name": 1, "quantity_mt": 1, "net_weight_mt": 1,
            "lifting_type": 1, "unloading_status": 1, "lifting_no": 1, "unloading_point_name": 1,
            "date_of_loading": 1, "date_of_unloading": 1, "vehicle_number": 1
        }}
    ]
    company_deliveries = await db.liftings.aggregate(company_deliveries_pipeline).to_list(500)
    
    # Aggregate company deliveries by product
    company_deliveries_by_product = {}
    for l in company_deliveries:
        pid = l.get("product_id") or "unknown"
        pname = l.get("product_name") or "Unknown"
        qty = l.get("net_weight_mt") or l.get("quantity_mt") or 0
        if pid not in company_deliveries_by_product:
            company_deliveries_by_product[pid] = {
                "product_name": pname, 
                "total_qty": 0, 
                "count": 0,
                "verified_qty": 0,
                "pending_qty": 0,
                "liftings": []
            }
        company_deliveries_by_product[pid]["total_qty"] += qty
        company_deliveries_by_product[pid]["count"] += 1
        if l.get("unloading_status") == "Verified":
            company_deliveries_by_product[pid]["verified_qty"] += qty
        else:
            company_deliveries_by_product[pid]["pending_qty"] += qty
        # Add lifting details for ledger
        company_deliveries_by_product[pid]["liftings"].append({
            "lifting_no": l.get("lifting_no"),
            "company_name": l.get("unloading_point_name"),
            "quantity": qty,
            "lifting_type": l.get("lifting_type"),
            "status": l.get("unloading_status"),
            "date_of_loading": l.get("date_of_loading"),
            "date_of_unloading": l.get("date_of_unloading"),
            "vehicle_number": l.get("vehicle_number")
        })
    
    # Pending verification breakdown by product
    pending_liftings = await db.liftings.find(
        {"unloading_status": "Pending"},
        {"_id": 0, "product_id": 1, "product_name": 1, "quantity_mt": 1, "lifting_type": 1}
    ).to_list(1000)
    
    pending_by_product = {}
    for l in pending_liftings:
        pid = l.get("product_id") or "unknown"
        pname = l.get("product_name") or "Unknown"
        qty = l.get("quantity_mt") or 0
        if pid not in pending_by_product:
            pending_by_product[pid] = {"product_name": pname, "total_qty": 0, "count": 0}
        pending_by_product[pid]["total_qty"] += qty
        pending_by_product[pid]["count"] += 1
    
    # Build liftings summary with X/Y format (lifted/total DO qty)
    liftings_product_wise = []
    all_product_ids = set(list(do_qty_by_product.keys()) + list(primary_by_product.keys()) + list(secondary_by_product.keys()))
    for pid in all_product_ids:
        do_info = do_qty_by_product.get(pid, {"product_name": "Unknown", "total_qty": 0})
        primary_info = primary_by_product.get(pid, {"product_name": do_info["product_name"], "lifted_qty": 0, "to_depot": 0, "to_client": 0})
        secondary_info = secondary_by_product.get(pid, {"product_name": do_info["product_name"], "lifted_qty": 0, "to_company": 0, "to_depot": 0})
        
        liftings_product_wise.append({
            "product_id": pid,
            "product_name": primary_info["product_name"] or secondary_info["product_name"] or do_info["product_name"],
            "do_total_qty": round(do_info["total_qty"], 2),
            "primary_lifted": round(primary_info["lifted_qty"], 2),
            "primary_to_depot": round(primary_info["to_depot"], 2),
            "primary_to_client": round(primary_info["to_client"], 2),
            "secondary_lifted": round(secondary_info["lifted_qty"], 2),
            "secondary_to_company": round(secondary_info["to_company"], 2),
            "secondary_to_depot": round(secondary_info["to_depot"], 2)
        })

    # Get master product names from Products collection
    all_products = await db.products.find(product_filter, {"_id": 0, "id": 1, "product_name": 1}).to_list(500)
    product_name_lookup = {p.get("id"): p.get("product_name") for p in all_products}
    
    # Include ALL products from the lookup, not just those with activity
    product_metric_ids = set(product_name_lookup.keys())
    product_metrics = []
    for pid in product_metric_ids:
        do_info = do_remaining_by_product.get(pid, {"product_name": "", "remaining_qty": 0})
        stock_info = available_stock_by_product.get(pid, {"product_name": "", "available_stock": 0})
        po_info = po_remaining_by_product.get(pid, {"product_name": "", "remaining_qty": 0})
        dispatch_info = dispatch_by_product.get(pid, {"product_name": "", "today": 0, "yesterday": 0, "day_before_yesterday": 0})

        # Use product lookup as primary source, fallback to aggregation results
        product_name = (
            product_name_lookup.get(pid) or
            do_info.get("product_name") or
            stock_info.get("product_name") or
            po_info.get("product_name") or
            dispatch_info.get("product_name") or
            "Unknown"
        )

        product_metrics.append({
            "product_id": pid,
            "product_name": product_name,
            "remaining_do_qty": round(do_info["remaining_qty"], 2),
            "available_stock_qty": round(stock_info["available_stock"], 2),
            "remaining_po_qty": round(po_info["remaining_qty"], 2),
            "dispatch_today_qty": round(dispatch_info["today"], 2),
            "dispatch_yesterday_qty": round(dispatch_info["yesterday"], 2),
            "dispatch_day_before_yesterday_qty": round(dispatch_info["day_before_yesterday"], 2),
            "stock_by_depot": stock_by_depot_by_product.get(pid, [])
        })

    # Add lightweight debug info to help diagnose empty dashboards
    try:
        debug_info = {
            "product_filter": str(product_filter),
            "assigned_products": list(current_user.get("assigned_products") or []),
            "product_metrics_count": len(product_metrics),
            "dispatch_products_count": len(dispatch_by_product),
            "dispatch_product_ids": list(dispatch_by_product.keys())[:10]
        }
    except Exception:
        debug_info = {"note": "unable to build debug info"}

    return {
        "counts": {
            "companies": companies_count,
            "users": users_count,
            "transporters": transporters_count,
            "trucks": trucks_count,
            "products": products_count,
            "depots": depots_count,
            "delivery_orders": orders_count,
            "liftings": liftings_count
        },
        "orders_by_status": {
            "open": open_orders,
            "in_progress": in_progress,
            "completed": completed
        },
        "liftings_by_status": {
            "pending": pending_verification,
            "verified": verified
        },
        # Detailed breakdowns
        "open_delivery_orders": open_dos,
        "liftings_product_wise": liftings_product_wise,
        "product_metrics": product_metrics,
        "pending_by_product": [
            {"product_id": k, **v} for k, v in pending_by_product.items()
        ],
        "primary_summary": {
            "total_count": len(primary_liftings),
            "total_qty": round(sum(l.get("quantity_mt", 0) for l in primary_liftings), 2),
            "to_depot_count": len(primary_to_depot),
            "to_depot_qty": round(sum(l.get("quantity_mt", 0) for l in primary_to_depot), 2),
            "to_client_count": len(primary_to_client),
            "to_client_qty": round(sum(l.get("quantity_mt", 0) for l in primary_to_client), 2)
        },
        "secondary_summary": {
            "total_count": len(secondary_liftings),
            "total_qty": round(sum(l.get("quantity_mt", 0) for l in secondary_liftings), 2),
            "to_company_count": len(secondary_to_company),
            "to_company_qty": round(sum(l.get("quantity_mt", 0) for l in secondary_to_company), 2),
            "to_depot_count": len(secondary_to_depot),
            "to_depot_qty": round(sum(l.get("quantity_mt", 0) for l in secondary_to_depot), 2)
        },
        # Company deliveries breakdown (product-wise with ledger)
        "company_deliveries": [
            {"product_id": k, **v} for k, v in company_deliveries_by_product.items()
        ],
        "company_deliveries_total": {
            "total_qty": round(sum(l.get("quantity_mt") or l.get("net_weight_mt", 0) for l in company_deliveries), 2),
            "total_count": len(company_deliveries),
            "verified_count": len([l for l in company_deliveries if l.get("unloading_status") == "Verified"]),
            "pending_count": len([l for l in company_deliveries if l.get("unloading_status") != "Verified"])
        }
    }

# ============ ROOT ============

@api_router.get("/")
async def root():
    return {"message": "LogiTrack Pro API v2.1 - InfoEIGHT"}

# Import and include all routers from routes package
from routes import (
    reports_router,
    companies_router,
    transporters_router,
    trucks_router,
    products_router,
    railway_sidings_router,
    railway_zones_router,
    depots_router,
    delivery_orders_router,
    liftings_router,
    permissions_router,
    product_access_router,
    depot_access_router,
    purchase_orders_router,
    pickups_router
    ,verified_trucks_router
)

# Include all modular routers
api_router.include_router(reports_router)
api_router.include_router(companies_router)
api_router.include_router(transporters_router)
api_router.include_router(trucks_router)
api_router.include_router(products_router)
api_router.include_router(railway_sidings_router)
api_router.include_router(railway_zones_router)
api_router.include_router(depots_router)
api_router.include_router(delivery_orders_router)
api_router.include_router(liftings_router)
api_router.include_router(permissions_router)
api_router.include_router(product_access_router)
api_router.include_router(depot_access_router)
api_router.include_router(purchase_orders_router)
api_router.include_router(pickups_router)
api_router.include_router(verified_trucks_router)

# Include router and middleware
app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

# Master Admin Configuration
MASTER_ADMIN_MOBILE = os.environ.get('MASTER_ADMIN_MOBILE')
MASTER_ADMIN_COUNTRY_CODE = os.environ.get('MASTER_ADMIN_COUNTRY_CODE', '91')
MASTER_ADMIN_PASSWORD = os.environ.get('MASTER_ADMIN_PASSWORD')
MASTER_ADMIN_NAME = os.environ.get('MASTER_ADMIN_NAME', 'Master Admin')
MASTER_ADMIN_EMAIL = os.environ.get('MASTER_ADMIN_EMAIL', 'admin@logitrackpro.com')

@app.on_event("startup")
async def seed_master_admin():
    """Seed Master Admin on startup if not exists, or update if credentials changed"""
    if not MASTER_ADMIN_MOBILE or not MASTER_ADMIN_PASSWORD:
        logger.warning("Master Admin credentials not configured in environment variables")
        return
        
    full_mobile = f"{MASTER_ADMIN_COUNTRY_CODE}{MASTER_ADMIN_MOBILE}"
    
    # Check if any master admin already exists
    existing = await db.users.find_one({"is_master_admin": True})
    
    if existing:
        # Update existing master admin with new credentials
        await db.users.update_one(
            {"is_master_admin": True},
            {"$set": {
                "mobile": full_mobile,
                "country_code": MASTER_ADMIN_COUNTRY_CODE,
                "password": hash_password(MASTER_ADMIN_PASSWORD),
                "email": MASTER_ADMIN_EMAIL,
                "name": MASTER_ADMIN_NAME
            }}
        )
        logger.info(f"Master Admin updated: +{full_mobile}")
        return
    
    # Create master admin
    master_admin = {
        "id": str(uuid.uuid4()),
        "name": MASTER_ADMIN_NAME,
        "mobile": full_mobile,
        "country_code": MASTER_ADMIN_COUNTRY_CODE,
        "password": hash_password(MASTER_ADMIN_PASSWORD),
        "password_set": True,
        "role": "Management",
        "email": MASTER_ADMIN_EMAIL,
        "depot_id": None,
        "otp_verified": True,
        "is_master_admin": True,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db.users.insert_one(master_admin)
    logger.info(f"Master Admin created: +{full_mobile} / {MASTER_ADMIN_PASSWORD}")

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()