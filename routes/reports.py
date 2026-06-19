"""Reports and export routes"""
from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from typing import Optional, List
from datetime import datetime, timezone
import io

from database import db
from auth_utils import get_current_user, require_permission

router = APIRouter(tags=["Reports"])

@router.get("/reports/liftings/date-wise")
async def get_datewise_lifting_report(
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    product_id: Optional[str] = None,
    company_id: Optional[str] = None,
    depot_id: Optional[str] = None,
    transport_mode: Optional[str] = None,
    lifting_type: Optional[str] = None,
    page: int = Query(1, ge=1, description="Page number"),
    page_size: int = Query(500, ge=10, le=1000, description="Items per page"),
    current_user: dict = Depends(require_permission("Lifting Reports"))
):
    """Get date-wise lifting report with filters and pagination"""
    
    # Build query
    query = {}
    
    if date_from or date_to:
        date_query = {}
        if date_from:
            date_query["$gte"] = date_from
        if date_to:
            date_query["$lte"] = date_to + "T23:59:59"  # Include full day
        if date_query:
            query["date_of_loading"] = date_query
    
    if product_id:
        query["product_id"] = product_id
    
    if company_id:
        query["loading_point_id"] = company_id
    
    if depot_id:
        query["unloading_point_id"] = depot_id
    
    if transport_mode:
        query["transport_mode"] = transport_mode
    
    if lifting_type:
        query["lifting_type"] = lifting_type
    
    # Get total count for pagination
    lifting_count = await db.liftings.count_documents(query)

    pickup_query = {"status": {"$in": ["verified", "weightment_done", "final_verified"]}}
    if date_from or date_to:
        pickup_date_query = {}
        if date_from:
            pickup_date_query["$gte"] = date_from
        if date_to:
            pickup_date_query["$lte"] = date_to + "T23:59:59"
        if pickup_date_query:
            pickup_query["verified_at"] = pickup_date_query
    if product_id:
        pickup_query["product_id"] = product_id
    if depot_id:
        pickup_query["depot_id"] = depot_id

    pickup_count = await db.pickups.count_documents(pickup_query)
    total_count = lifting_count + pickup_count

    # Calculate pagination
    skip = (page - 1) * page_size
    total_pages = (total_count + page_size - 1) // page_size

    # Get liftings with pagination
    liftings = await db.liftings.find(query, {"_id": 0}).sort("date_of_loading", -1).skip(skip).limit(page_size).to_list(page_size)

    # Add verified pickups as stock movement rows so reports match inventory ledger
    pickups = await db.pickups.find(pickup_query, {"_id": 0}).sort("verified_at", -1).to_list(1000)
    for pickup in pickups:
        liftings.append({
            **pickup,
            "lifting_type": "Pickup",
            "lifting_no": pickup.get("purchase_order_no") or pickup.get("purchase_order_id") or pickup.get("id"),
            "quantity_mt": pickup.get("loaded_weight_mt") or pickup.get("weight_mt", 0),
            "transport_mode": "Road",
            "vehicle_number": pickup.get("truck_number"),
            "loading_point_name": pickup.get("depot_name"),
            "unloading_point_name": pickup.get("purchase_order_company_name") or pickup.get("company_name"),
            "unloading_status": "Verified",
            "date_of_loading": pickup.get("verified_at") or pickup.get("date")
        })

    # Group by date
    date_groups = {}
    for lifting in liftings:
        date_value = lifting.get("date_of_loading") or "Unknown"
        date_key = date_value[:10] if len(date_value) >= 10 else date_value  # YYYY-MM-DD
        if date_key not in date_groups:
            date_groups[date_key] = {
                "date": date_key,
                "total_liftings": 0,
                "total_quantity_mt": 0,
                "total_net_weight_mt": 0,
                "by_product": {},
                "by_transport_mode": {"Road": 0, "Railway": 0},
                "liftings": []
            }
        
        group = date_groups[date_key]
        group["total_liftings"] += 1
        group["total_quantity_mt"] += lifting.get("quantity_mt", 0) or 0
        group["total_net_weight_mt"] += lifting.get("net_weight_mt", 0) or 0
        
        # Group by product
        product_name = lifting.get("product_name", "Unknown")
        if product_name not in group["by_product"]:
            group["by_product"][product_name] = {"count": 0, "quantity_mt": 0}
        group["by_product"][product_name]["count"] += 1
        group["by_product"][product_name]["quantity_mt"] += lifting.get("quantity_mt", 0) or 0
        
        # Group by transport mode
        mode = lifting.get("transport_mode", "Road")
        group["by_transport_mode"][mode] = group["by_transport_mode"].get(mode, 0) + 1
        
        group["liftings"].append(lifting)
    
    # Convert to sorted list
    result = sorted(date_groups.values(), key=lambda x: x["date"], reverse=True)
    
    # Calculate summary
    summary = {
        "total_dates": len(result),
        "total_liftings": sum(g["total_liftings"] for g in result),
        "total_quantity_mt": sum(g["total_quantity_mt"] for g in result),
        "total_net_weight_mt": sum(g["total_net_weight_mt"] for g in result),
        "by_transport_mode": {
            "Road": sum(g["by_transport_mode"].get("Road", 0) for g in result),
            "Railway": sum(g["by_transport_mode"].get("Railway", 0) for g in result)
        }
    }
    
    return {
        "summary": summary,
        "data": result,
        "pagination": {
            "page": page,
            "page_size": page_size,
            "total_count": total_count,
            "total_pages": total_pages,
            "has_next": page < total_pages,
            "has_prev": page > 1
        },
        "filters_applied": {
            "date_from": date_from,
            "date_to": date_to,
            "product_id": product_id,
            "company_id": company_id,
            "depot_id": depot_id,
            "transport_mode": transport_mode,
            "lifting_type": lifting_type
        }
    }


@router.get("/reports/liftings/export")
async def export_datewise_lifting_report(
    format: str = "excel",
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    product_id: Optional[str] = None,
    company_id: Optional[str] = None,
    depot_id: Optional[str] = None,
    transport_mode: Optional[str] = None,
    lifting_type: Optional[str] = None,
    current_user: dict = Depends(require_permission("Lifting Reports"))
):
    """Export date-wise lifting report to Excel"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    
    # Build query
    query = {}
    
    if date_from or date_to:
        date_query = {}
        if date_from:
            date_query["$gte"] = date_from
        if date_to:
            date_query["$lte"] = date_to + "T23:59:59"
        if date_query:
            query["date_of_loading"] = date_query
    
    if product_id:
        query["product_id"] = product_id
    if company_id:
        query["loading_point_id"] = company_id
    if depot_id:
        query["unloading_point_id"] = depot_id
    if transport_mode:
        query["transport_mode"] = transport_mode
    if lifting_type:
        query["lifting_type"] = lifting_type
    
    # Export with reasonable limit (max 5000 records for Excel export)
    liftings = await db.liftings.find(query, {"_id": 0}).sort("date_of_loading", -1).to_list(5000)
    
    if format == "excel":
        # Add verified pickups into export too so export matches the report rows
        pickup_query = {"status": {"$in": ["verified", "weightment_done", "final_verified"]}}
        if date_from or date_to:
            pickup_date_query = {}
            if date_from:
                pickup_date_query["$gte"] = date_from
            if date_to:
                pickup_date_query["$lte"] = date_to + "T23:59:59"
            if pickup_date_query:
                pickup_query["verified_at"] = pickup_date_query
        if product_id:
            pickup_query["product_id"] = product_id
        if depot_id:
            pickup_query["depot_id"] = depot_id

        pickups = await db.pickups.find(pickup_query, {"_id": 0}).sort("verified_at", -1).to_list(1000)
        pickup_records = []
        for pickup in pickups:
            pickup_records.append({
                **pickup,
                "lifting_type": "Pickup",
                "lifting_no": pickup.get("purchase_order_no") or pickup.get("purchase_order_id") or pickup.get("id"),
                "quantity_mt": pickup.get("loaded_weight_mt") or pickup.get("weight_mt", 0),
                "transport_mode": "Road",
                "vehicle_number": pickup.get("truck_number"),
                "loading_point_name": pickup.get("depot_name"),
                "unloading_point_name": pickup.get("purchase_order_company_name") or pickup.get("company_name"),
                "unloading_status": "Verified",
                "date_of_loading": pickup.get("verified_at") or pickup.get("date")
            })

        liftings.extend(pickup_records)
        liftings.sort(key=lambda x: x.get("date_of_loading") or "", reverse=True)

        wb = Workbook()
        ws = wb.active
        ws.title = "Date-wise Lifting Report"

        # Header styling
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Title row
        ws.merge_cells('A1:N1')
        title_cell = ws.cell(row=1, column=1, value="Date-wise Lifting Report")
        title_cell.font = Font(bold=True, size=14)
        title_cell.alignment = Alignment(horizontal='center')
        
        # Filter info
        filter_text = f"Period: {date_from or 'All'} to {date_to or 'All'}"
        ws.merge_cells('A2:N2')
        ws.cell(row=2, column=1, value=filter_text).alignment = Alignment(horizontal='center')
        
        # Headers
        headers = [
            "Date", "Lifting No", "Type", "Transport Mode", "Product", "Product Code",
            "Quantity (MT)", "Vehicle/Rake", "From", "To", "Status", 
            "Net Weight (MT)", "Driver", "Verified By"
        ]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border
        
        # Data rows
        row = 5
        for lifting in liftings:
            ws.cell(row=row, column=1, value=lifting.get("date_of_loading", "")[:10] if lifting.get("date_of_loading") else "")
            ws.cell(row=row, column=2, value=lifting.get("lifting_no", ""))
            ws.cell(row=row, column=3, value=lifting.get("lifting_type", ""))
            ws.cell(row=row, column=4, value=lifting.get("transport_mode", "Road"))
            ws.cell(row=row, column=5, value=lifting.get("product_name", ""))
            ws.cell(row=row, column=6, value=lifting.get("product_code", ""))
            ws.cell(row=row, column=7, value=lifting.get("quantity_mt", 0))
            
            # Vehicle or Railway info
            if lifting.get("transport_mode") == "Railway":
                ws.cell(row=row, column=8, value=lifting.get("loading_siding_name", ""))
            else:
                ws.cell(row=row, column=8, value=lifting.get("vehicle_number", ""))
            
            ws.cell(row=row, column=9, value=lifting.get("loading_point_name", ""))
            ws.cell(row=row, column=10, value=lifting.get("unloading_point_name", ""))
            ws.cell(row=row, column=11, value=lifting.get("unloading_status", ""))
            ws.cell(row=row, column=12, value=lifting.get("net_weight_mt", 0))
            ws.cell(row=row, column=13, value=lifting.get("driver_name", ""))
            ws.cell(row=row, column=14, value=lifting.get("verified_by_name", ""))
            
            for col in range(1, 15):
                ws.cell(row=row, column=col).border = thin_border
            row += 1
        
        # Summary row
        row += 1
        ws.cell(row=row, column=1, value="TOTAL").font = Font(bold=True)
        ws.cell(row=row, column=7, value=sum(item.get("quantity_mt", 0) or 0 for item in liftings)).font = Font(bold=True)
        ws.cell(row=row, column=12, value=sum(item.get("net_weight_mt", 0) or 0 for item in liftings)).font = Font(bold=True)
        
        # Auto-width columns
        for col in ws.columns:
            max_length = 0
            # Skip merged cells to avoid AttributeError
            if hasattr(col[0], 'column_letter'):
                column = col[0].column_letter
                for cell in col:
                    try:
                        if hasattr(cell, 'value') and len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except Exception:
                        pass
                adjusted_width = min(max_length + 2, 30)
                ws.column_dimensions[column].width = adjusted_width
        
        # Save to buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        filename = f"datewise_lifting_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        return StreamingResponse(
            buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    
    return {"error": "Unsupported format"}


@router.get("/reports/summary")
async def get_reports_summary(
    current_user: dict = Depends(get_current_user)
):
    """Get summary statistics for reports dashboard"""
    
    today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    
    # Today's liftings
    today_liftings = await db.liftings.count_documents({
        "date_of_loading": {"$regex": f"^{today}"}
    })
    
    # Total liftings
    total_liftings = await db.liftings.count_documents({})
    
    # Pending verifications
    pending_verifications = await db.liftings.count_documents({
        "unloading_status": "Pending"
    })
    
    # Open delivery orders
    open_orders = await db.delivery_orders.count_documents({
        "status": "Open"
    })
    
    # Products count
    products_count = await db.products.count_documents({})
    
    # Active depots
    depots_count = await db.depots.count_documents({})
    
    return {
        "today_liftings": today_liftings,
        "total_liftings": total_liftings,
        "pending_verifications": pending_verifications,
        "open_delivery_orders": open_orders,
        "products_count": products_count,
        "depots_count": depots_count
    }
